import re
from typing import Any, Dict, List, Tuple, Optional
from openpyxl.worksheet.worksheet import Worksheet

from utils import norm, extract_imeis, _luhn_ok_imei

HEADERS = [
    "imei","Bulunma","Tck/Vkn","Belge Türü","Belge Tarihi","Belge No","Alınan Kişi","Borç Tutar",
    "Marka","MODEL",
    "SATIŞ TARİHi","ALICI ADI SOYADI","SATIŞ BEDELİ","KDV TUTARI","SATIŞ BELGESİNİN NUMARASI",
    "ALICI KİMLİK TÜRÜ","ALICI KİMLİK NO",
    "Sütun1","ALIŞ BELGELERİ TÜRÜ","DURUMU",
    "ALIŞ KDV","SATIŞ KDV","SINIF","GEREKÇE",
]


def ensure_len(vals: List[Any]) -> List[Any]:
    if len(vals) < len(HEADERS):
        vals = vals + [""] * (len(HEADERS) - len(vals))
    elif len(vals) > len(HEADERS):
        vals = vals[:len(HEADERS)]
    return vals


GP_KW = {
    "imei": ["imei", "seri", "serial", "serı", "serino", "seri no", "seri-no"],
    "tarih": ["tarih", "alış tarihi", "alis tarihi", "işlem tarihi", "islem tarihi", "gp tarihi"],
    "bedel": ["tutar", "bedel", "fiyat", "ödenen", "odenen", "ucret", "ücret"],
    "ad": ["adı", "ad soyad", "adsoyad", "satıcı", "satici", "müsteri", "musteri", "alan", "veren"],
    "sube": ["şube", "sube", "magaza", "mağaza"],
    "aciklama": ["açıklama", "aciklama", "not", "ürün", "urun", "cihaz", "model"],
}


def _find_col_idx(headers_row: List[Any], keys: List[str]) -> Optional[int]:
    for i, h in enumerate(headers_row):
        hh = norm(h).lower()
        for k in keys:
            if k in hh:
                return i
    return None


def _scan_header(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    for r, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(12, ws.max_row), values_only=True),
        start=1,
    ):
        idx_imei = _find_col_idx(row, GP_KW["imei"])
        if idx_imei is not None:
            cols = {
                "imei": idx_imei,
                "tarih": _find_col_idx(row, GP_KW["tarih"]) or -1,
                "bedel": _find_col_idx(row, GP_KW["bedel"]) or -1,
                "ad": _find_col_idx(row, GP_KW["ad"]) or -1,
                "sube": _find_col_idx(row, GP_KW["sube"]) or -1,
                "aciklama": _find_col_idx(row, GP_KW["aciklama"]) or -1,
            }
            return r, cols
    cols = {"imei": 0, "tarih": -1, "bedel": -1, "ad": -1, "sube": -1, "aciklama": -1}
    return 1, cols


def parse_gp_workbook(wb, log) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        head_row, cols = _scan_header(ws)
        start = head_row + 1
        found_rows = 0
        for row in ws.iter_rows(min_row=start, values_only=True):
            text_row = " | ".join([norm(v) for v in row])
            imeis = extract_imeis(text_row)
            if not imeis and cols["imei"] >= 0:
                v = norm(row[cols["imei"]]) if cols["imei"] < len(row) else ""
                if re.fullmatch(r"\d{15}", v) and _luhn_ok_imei(v):
                    imeis = [v]
            if not imeis:
                continue
            for im in imeis:
                item = {
                    "imei": im,
                    "tarih": norm(row[cols["tarih"]]) if cols["tarih"] >= 0 else "",
                    "bedel": norm(row[cols["bedel"]]) if cols["bedel"] >= 0 else "",
                    "ad": norm(row[cols["ad"]]) if cols["ad"] >= 0 else "",
                    "sube": norm(row[cols["sube"]]) if cols["sube"] >= 0 else "",
                    "aciklama": norm(row[cols["aciklama"]]) if cols["aciklama"] >= 0 else "",
                }
                out.append(item)
                found_rows += 1
        log(f"[GP] Sayfa '{ws.title}': {found_rows} satır/IMEI çıkarıldı.")
    return out
