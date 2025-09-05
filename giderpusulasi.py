import re
from typing import Any, List, Tuple

TR_MAP = str.maketrans("çğıöşüÇĞİÖŞÜ","cgiosuCGIOSU")

def norm(s: Any) -> str:
    return ("" if s is None else str(s)).strip()

def nlow(s: Any) -> str:
    return norm(s).lower().translate(TR_MAP)

def nup(s: Any) -> str:
    return norm(s).upper().translate(TR_MAP)

IMEI_RE_STRICT = re.compile(r'(?<!\d)\d{15}(?!\d)')


def _luhn_ok_imei(s15: str) -> bool:
    if len(s15) != 15 or not s15.isdigit():
        return False
    total = 0
    for i, ch in enumerate(s15):
        n = int(ch)
        if i % 2 == 1:
            n *= 2
            if n > 9:
                n -= 9
        total += n
    return (total % 10) == 0

def extract_imeis(text: str) -> List[str]:
    t = norm(text)
    if not t:
        return []
    out = set()
    for m in IMEI_RE_STRICT.finditer(t):
        s = m.group(0)
        if _luhn_ok_imei(s):
            out.add(s)
    return sorted(out)


def safe_filename(name: str) -> str:
    s = norm(name)
    s = re.sub(r'[\\/*?:"<>|]', "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    return (s or "DOSYA")[:120]


def _cp(p: str) -> re.Pattern:
    return re.compile(p, re.I)

_BRAND_PATTERNS: List[Tuple[str, re.Pattern]] = [
    ("APPLE",   _cp(r"\bAPPLE\b|\bIPHONE?\b|\bAPLE\b|\bI ?PHONE\b")),
    ("SAMSUNG", _cp(r"\bSAMSUNG\b|\bGALAXY\b|\bSM[-\s]")),
    ("XIAOMI",  _cp(r"\bXIAOM[Iİ]\b|\bREDMI\b|\bPOCO\b|\bMI[-\s]")),
    ("HUAWEI",  _cp(r"\bHUAWEI\b|\bHUAWE\b|\bP\d{2}\b|\bMATE\b(?!.*HONOR)")),
    ("HONOR",   _cp(r"\bHONOR\b")),
    ("OPPO",    _cp(r"\bOPPO\b")),
    ("REALME",  _cp(r"\bREALME\b")),
    ("VIVO",    _cp(r"\bVIVO\b")),
    ("TECNO",   _cp(r"\bTECNO\b")),
    ("NOKIA",   _cp(r"\bNOKIA\b")),
    ("CASPER",  _cp(r"\bCASPER\b")),
    ("GENERAL MOBILE", _cp(r"\bGENERAL\s*MOBILE\b|\bGM\s?\d+\b")),
    ("INFINIX", _cp(r"\bINFINIX\b")),
    ("REEDER",  _cp(r"\bREEDER\b")),
]

def brand_from_text(t: str) -> str:
    U = nup(t)
    for brand, pat in _BRAND_PATTERNS:
        if pat.search(U):
            if brand == "HUAWEI" and "HONOR" in U:
                continue
            return brand
    return "Bilinmeyen"

KEY_2EL  = re.compile(r"\b(2\.?\s*EL|İKİNCİ\s*EL|IKINCI\s*EL|SECOND\s*HAND)\b", re.I)
KEY_REF  = re.compile(r"\b(YENİLENMİŞ|YENILENMIS|REFURB)\b", re.I)
DOCNO_RE = re.compile(r"\bE(?:AR|FR)\d{13}\b", re.I)
import logging
import tkinter as tk

def setup_logger(widget: tk.Text = None, logfile: str = "app.log") -> logging.Logger:
    logger = logging.getLogger("giderpusulasi")
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    fh = logging.FileHandler(logfile, encoding="utf-8")
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    if widget is not None:
        class TextHandler(logging.Handler):
            def __init__(self, text_widget: tk.Text):
                super().__init__()
                self.widget = text_widget
            def emit(self, record: logging.LogRecord):
                msg = self.format(record)
                self.widget.insert(tk.END, msg + "\n")
                self.widget.see(tk.END)
        th = TextHandler(widget)
        th.setFormatter(formatter)
        logger.addHandler(th)
    return logger
import re
from typing import Any, Dict, List, Tuple, Optional
from openpyxl.worksheet.worksheet import Worksheet


HEADERS = [
    "imei","Bulunma","Tck/Vkn","Belge Türü","Belge Tarihi","Belge No","Alınan Kişi","Borç Tutar",
    "Marka","MODEL",
    "SATIŞ TARİHi","ALICI ADI SOYADI","SATIŞ BEDELİ","KDV TUTARI","SATIŞ BELGESİNİN NUMARASI",
    "ALICI KİMLİK TÜRÜ","ALICI KİMLİK NO",
    "Sütun1","ALIŞ BELGELERİ TÜRÜ","DURUMU",
    "ALIŞ KDV","SATIŞ KDV","SINIF","GEREKÇE",
]


def ensure_len(vals: List[Any]) -> List[Any]:
    if len(vals) < len(HEADERS):
        vals = vals + [""] * (len(HEADERS) - len(vals))
    elif len(vals) > len(HEADERS):
        vals = vals[:len(HEADERS)]
    return vals


GP_KW = {
    "imei": ["imei", "seri", "serial", "serı", "serino", "seri no", "seri-no"],
    "tarih": ["tarih", "alış tarihi", "alis tarihi", "işlem tarihi", "islem tarihi", "gp tarihi"],
    "bedel": ["tutar", "bedel", "fiyat", "ödenen", "odenen", "ucret", "ücret"],
    "ad": ["adı", "ad soyad", "adsoyad", "satıcı", "satici", "müsteri", "musteri", "alan", "veren"],
    "sube": ["şube", "sube", "magaza", "mağaza"],
    "aciklama": ["açıklama", "aciklama", "not", "ürün", "urun", "cihaz", "model"],
}


def _find_col_idx(headers_row: List[Any], keys: List[str]) -> Optional[int]:
    for i, h in enumerate(headers_row):
        hh = norm(h).lower()
        for k in keys:
            if k in hh:
                return i
    return None


def _scan_header(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    for r, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(12, ws.max_row), values_only=True),
        start=1,
    ):
        idx_imei = _find_col_idx(row, GP_KW["imei"])
        if idx_imei is not None:
            cols = {
                "imei": idx_imei,
                "tarih": _find_col_idx(row, GP_KW["tarih"]) or -1,
                "bedel": _find_col_idx(row, GP_KW["bedel"]) or -1,
                "ad": _find_col_idx(row, GP_KW["ad"]) or -1,
                "sube": _find_col_idx(row, GP_KW["sube"]) or -1,
                "aciklama": _find_col_idx(row, GP_KW["aciklama"]) or -1,
            }
            return r, cols
    cols = {"imei": 0, "tarih": -1, "bedel": -1, "ad": -1, "sube": -1, "aciklama": -1}
    return 1, cols


def parse_gp_workbook(wb, log) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        head_row, cols = _scan_header(ws)
        start = head_row + 1
        found_rows = 0
        for row in ws.iter_rows(min_row=start, values_only=True):
            text_row = " | ".join([norm(v) for v in row])
            imeis = extract_imeis(text_row)
            if not imeis and cols["imei"] >= 0:
                v = norm(row[cols["imei"]]) if cols["imei"] < len(row) else ""
                if re.fullmatch(r"\d{15}", v) and _luhn_ok_imei(v):
                    imeis = [v]
            if not imeis:
                continue
            for im in imeis:
                item = {
                    "imei": im,
                    "tarih": norm(row[cols["tarih"]]) if cols["tarih"] >= 0 else "",
                    "bedel": norm(row[cols["bedel"]]) if cols["bedel"] >= 0 else "",
                    "ad": norm(row[cols["ad"]]) if cols["ad"] >= 0 else "",
                    "sube": norm(row[cols["sube"]]) if cols["sube"] >= 0 else "",
                    "aciklama": norm(row[cols["aciklama"]]) if cols["aciklama"] >= 0 else "",
                }
                out.append(item)
                found_rows += 1
        log(f"[GP] Sayfa '{ws.title}': {found_rows} satır/IMEI çıkarıldı.")
    return out
import math
import threading
from typing import Any, Dict, List, Optional

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# NES API endpoints
EINV_IN_LIST   = "https://api.nes.com.tr/einvoice/v1/incoming/invoices"
EINV_OUT_LIST  = "https://api.nes.com.tr/einvoice/v1/outgoing/invoices"
EARCH_OUT_LIST = "https://api.nes.com.tr/earchive/v1/invoices"

EINV_IN_DOC    = "https://api.nes.com.tr/einvoice/v1/incoming/invoices/{id}"
EINV_OUT_DOC   = "https://api.nes.com.tr/einvoice/v1/outgoing/invoices/{id}"
EARCH_OUT_DOC  = "https://api.nes.com.tr/earchive/v1/invoices/{id}"

PAGE_SIZE = 50

# HTTP session configuration
TIMEOUT_CONNECT = 15
TIMEOUT_READ = 90
RETRIES = 4
BACKOFF = 0.6
SESSION: Optional[requests.Session] = None


def make_session(retries: int = None, backoff: float = None) -> requests.Session:
    r = retries if retries is not None else RETRIES
    b = backoff if backoff is not None else BACKOFF
    s = requests.Session()
    retry = Retry(
        total=r, connect=r, read=r, status=r,
        backoff_factor=b,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET", "POST"]),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=100, pool_maxsize=100)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s


def get_session() -> requests.Session:
    global SESSION
    if SESSION is None:
        SESSION = make_session()
    return SESSION


def headers(token: str, typ: str = "json") -> Dict[str, str]:
    h = {
        "Authorization": f"Bearer {token.strip()}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": "IMEI-NES-Client/10.3",
    }
    if typ == "xml":
        h["Accept"] = "application/xml"
    if typ == "pdf":
        h["Accept"] = "application/pdf"
    if typ == "bin":
        h["Accept"] = "*/*"
    return h


def http_get(
    url: str,
    *,
    token: Optional[str],
    typ: str = "json",
    params=None,
    log=None,
    stop_evt: Optional[threading.Event] = None,
):
    if stop_evt is not None and stop_evt.is_set():
        return None
    try:
        sess = get_session()
        hdrs = headers(token, typ) if token is not None else {"User-Agent": "IMEI-NES-Client/10.3"}
        r = sess.get(url, headers=hdrs, params=params, timeout=(TIMEOUT_CONNECT, TIMEOUT_READ))
        if r.status_code == 200:
            return r
        if log:
            log(f"[HTTP] Hata {r.status_code}: {url} → {r.text[:300]}")
        return None
    except requests.Timeout:
        if log:
            log(f"[HTTP] Zaman aşımı: {url}")
        return None
    except requests.RequestException as e:
        if log:
            log(f"[HTTP] İstek hatası: {e}")
        return None


def paged_list(
    url: str,
    token: str,
    start: str,
    end: str,
    log,
    stop_evt: threading.Event,
    *,
    archived: Optional[bool] = None,
    section_name: str = "NES",
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    page = 1
    total = None
    while True:
        if stop_evt.is_set():
            break
        params = {"sort": "createdAt desc", "page": page, "pageSize": PAGE_SIZE}
        if archived is not None:
            params["archived"] = "true" if archived else "false"
        if start:
            params["startDate"] = f"{start}T00:00:00+03:00"
        if end:
            params["endDate"] = f"{end}T23:59:59+03:00"
        r = http_get(url, token=token, typ="json", params=params, log=log, stop_evt=stop_evt)
        if r is None:
            log(f"[{section_name}] İstek başarısız/timeout. Döngü sonlandırıldı.")
            break
        try:
            data = r.json() or {}
        except ValueError:
            log(f"[{section_name}] JSON çözümlenemedi.")
            break
        if total is None:
            tc = data.get("totalCount") or 0
            total = max(1, math.ceil(tc / PAGE_SIZE)) if isinstance(tc, int) else 1
        batch = data.get("data") or data.get("invoices") or []
        if not batch:
            break
        log(f"[{section_name}] Sayfa {page}/{total} → {len(batch)} kayıt")
        out.extend(batch)
        if page >= total:
            break
        page += 1
    return out


def list_both_archived(
    url: str,
    token: str,
    start: str,
    end: str,
    log,
    stop_evt: threading.Event,
    section_name: str,
) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    log(f"[{section_name}] Arşivsiz çekiliyor...")
    items += paged_list(url, token, start, end, log, stop_evt, archived=False, section_name=section_name)
    log(f"[{section_name}] Arşivli çekiliyor...")
    items += paged_list(url, token, start, end, log, stop_evt, archived=True, section_name=section_name)
    seen, uniq = set(), []
    for m in items:
        mid = str(m.get("id") or "")
        if mid and mid not in seen:
            seen.add(mid)
            uniq.append(m)
    log(f"[{section_name}] Birleştirildi (tekil): {len(uniq)} kayıt")
    return uniq


def fetch_xml_by(url_tpl: str, token: str, inv_id: str) -> Optional[bytes]:
    r = http_get(f"{url_tpl.format(id=inv_id)}/xml", token=token, typ="xml")
    return r.content if r is not None else None


def fetch_pdf_by(url_tpl: str, token: str, inv_id: str) -> Optional[bytes]:
    r = http_get(f"{url_tpl.format(id=inv_id)}/pdf", token=token, typ="pdf")
    return r.content if r is not None else None
# -*- coding: utf-8 -*-
"""
IMEI → Alış + Satış Birleşik Rapor (NES + GP) – v10.3

YENİ (v10.3):
- 'Arama Opsiyonları' altındaki metin alanı kaldırıldı; yerine IMEI Listesi paneli eklendi.
- Kaynak Seçimi: Hepsi | Sadece IMEI | Sadece Fatura No | Sadece URL | Sadece Excel (GP).
- 'Listeler dışı da tara?' seçeneği eklendi (kapalıyken sadece verilen listelere bağlı çalışır).
- Satış taramalarında meta kısayol kaldırıldı; filtreler XML parse SONRASI uygulanır (yalnızca XML içinden arama).
- Interaktif tablo: sütun başlığına tıklayarak sıralama, metin filtresi, 'Eksik bilgili satırları göster' filtresi,
  çoklu seçim ve Ctrl+C ile tabloyu Excel'e kopyalama.
- İndirme aksiyonları ayrıldı: Seçili→PDF, Seçili→XML, Seçili→PDF+XML; sağ tık menüsü ve çift tıkla hızlı PDF indir.
- Log bölümü PanedWindow ile genişletilebilir.
- Varsayılan tarih aralığı: başlangıç 2025-01-01, bitiş bugün.

v10.2'den devralınanlar korunmuştur.
"""

import os, re, io, json, math, threading, csv
from datetime import date
from typing import Any, Dict, List, Optional, Tuple, Set
import xml.etree.ElementTree as ET

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext


# ====================== Sabitler ======================
DEFAULTS = {
    "api_token": "3B51B4C7C94FF977E42389915CFDA353F6DCE2BF6A2A82C033FBB0950B17CDE8",
    "download_dir": os.getcwd(),
    "out_name": r"C:\Users\siyah\OneDrive\Masaüstü\imei_rapor.xlsx",
    # IMEI dışı fatura (kargo/enerji/telekom vs) beyaz liste: ünvan regex’leri
    "whitelist_patterns": [
        r"\bYURTIC[IİÇ]|ARAS|MNG|S[ÜU]RAT|PTT|UPS|FEDEX|DHL\b",
        r"\bT[ÜU]RK ?TELEKOM|TTNET|TURKCELL|VODAFONE|LIFECELL|T[ÜU]RKSAT|SUPERONLINE\b",
        r"\bELEKTR[Iİ]K|GAZ|DO[GĞ]AL ?GAZ|ENERJ[İI]|PERAKENDE SAT[Iİ]Ş\b",
        r"\bKULE Y[ÖO]NET[İI]M\b|\bLORAS GAYR[İI]MENKUL\b",
    ],
    "timeout_connect": 15,
    "timeout_read": 90,
    "retries": 4,
    "backoff": 0.6,
}
DEFAULT_DATE_START = "2025-01-01"
def _today_str(): return date.today().strftime("%Y-%m-%d")

SETTINGS_FILE = "imei_beyanname_v10.json"

# ---------- TR karakter dönüşümü ----------

# ---------- Belge no desenleri ----------


# ====================== Ayarlar / Excel ======================

def load_settings() -> Dict[str,Any]:
    s = DEFAULTS.copy()
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f: s.update(json.load(f))
        except Exception: pass
    try:
        s["_whitelist_compiled"] = [re.compile(pat, re.I) for pat in s.get("whitelist_patterns",[])]
    except Exception:
        s["_whitelist_compiled"] = []
    return s
def save_settings(s: Dict[str,Any]):
    s2 = {k:v for k,v in s.items() if k != "_whitelist_compiled"}
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f: json.dump(s2, f, ensure_ascii=False, indent=2)

def write_excel(rows: List[List[Any]], out_path: str):
    wb = Workbook(); ws = wb.active; ws.title = "IMEI_RAPOR"
    ws.append(HEADERS)
    for r in rows: ws.append(ensure_len(r))
    for col in range(1, len(HEADERS)+1):
        mx = max((len(str(ws.cell(row=i, column=col).value or "")) for i in range(1, ws.max_row+1)), default=12)
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, mx+2), 60)
    wb.save(out_path)

# ---------- Beyaz Liste ----------
def is_whitelisted_supplier(name: str, settings: Dict[str,Any]) -> bool:
    if not name: return False
    U = nup(name)
    for pat in settings.get("_whitelist_compiled", []):
        if pat.search(U):
            return True
    return False

# ====================== GP (Gider Pusulası) – Okuyucu (aynı) ======================
def _norm_header(h: Any) -> str:
    return nlow(h).replace("  ", " ").strip()

_HEADER_ALIAS = {
    "imei": "imei", "imei/seri": "imei", "seri no": "imei", "serino": "imei",
    "tck/vkn": "Tck/Vkn", "tckn": "Tck/Vkn", "vkn": "Tck/Vkn", "tc kimlik": "Tck/Vkn",
    "belge turu": "Belge Türü", "belge tarihi": "Belge Tarihi", "belge no": "Belge No",
    "alinan kisi": "Alınan Kişi", "alınan kişi": "Alınan Kişi",
    "borc tutar": "Borç Tutar",
    "marka": "Marka", "model": "MODEL", "açıklama": "MODEL", "aciklama": "MODEL", "urun": "MODEL", "ürün": "MODEL",
    "satis tarihi": "SATIŞ TARİHi", "alici adi soyadi": "ALICI ADI SOYADI",
    "satis bedeli": "SATIŞ BEDELİ", "kdv tutari": "KDV TUTARI",
    "satis belgesinin numarasi": "SATIŞ BELGESİNİN NUMARASI",
    "alici kimlik turu": "ALICI KİMLİK TÜRÜ", "alici kimlik no": "ALICI KİMLİK NO",
    "sutun1": "Sütun1", "şube": "Sütun1", "sube": "Sütun1", "magaza":"Sütun1",
    "alis belgeleri turu": "ALIŞ BELGELERİ TÜRÜ", "durumu": "DURUMU",
}

def _build_header_map(ws) -> Optional[Dict[str, Any]]:
    for r in range(1, min(11, ws.max_row)+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        normed = [_norm_header(x) for x in row]
        col_to_name = {}; hit = 0; has_imei = False
        for idx, hh in enumerate(normed):
            if not hh: continue
            key = hh.replace("ı","i")
            key = re.sub(r"[^a-z0-9/ ]", "", key)
            name = _HEADER_ALIAS.get(key, None)
            if name:
                col_to_name[idx] = name; hit += 1
                if name == "imei": has_imei = True
        if has_imei and hit >= 5:
            return {"row": r, "cols": col_to_name}
    return None

def parse_gp_template_workbook(wb, log) -> List[List[Any]]:
    out_rows: List[List[Any]] = []
    for ws in wb.worksheets:
        hm = _build_header_map(ws)
        if not hm: continue
        hdr_row = hm["row"]; colmap = hm["cols"]
        taken = 0
        for r in range(hdr_row+1, ws.max_row+1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
            if not any(norm(v) for v in vals): continue
            row_dict = {name: "" for name in HEADERS}
            for cidx, name in colmap.items():
                if cidx < len(vals):
                    row_dict[name] = norm(vals[cidx])
            imei_val = row_dict.get("imei", "")
            if not re.fullmatch(r"\d{15}", imei_val or ""):
                merged = " | ".join([norm(v) for v in vals])
                m_all = extract_imeis(merged)
                if not m_all: continue
                row_dict["imei"] = m_all[0]

            if nup(row_dict.get("Belge Türü","")) in ("GMA","GIDER PUSULASI","GİDER PUSULASI"):
                row_dict["Belge Türü"] = "GİDER PUSULASI"
                row_dict["ALIŞ BELGELERİ TÜRÜ"] = row_dict.get("ALIŞ BELGELERİ TÜRÜ") or "Gider Pusulası"
            if not row_dict.get("Marka"):
                row_dict["Marka"] = brand_from_text(row_dict.get("MODEL",""))
            if any(row_dict.get(k) for k in ("SATIŞ TARİHi","ALICI ADI SOYADI","SATIŞ BEDELİ","SATIŞ BELGESİNİN NUMARASI")):
                row_dict["DURUMU"] = row_dict.get("DURUMU") or "Satılmış"
            else:
                row_dict["DURUMU"] = row_dict.get("DURUMU") or "Satılabilir"
            row_dict["Bulunma"] = row_dict.get("Bulunma") or "GP"

            out_rows.append([row_dict[h] for h in HEADERS])
            taken += 1
        log(f"[GP] Sayfa '{ws.title}': {taken} satır alındı.")
    return out_rows

# ====================== GUI ======================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("IMEI → Alış + Satış Birleşik Rapor (NES + GP) – v10.3")
        self.geometry("1600x1000")
        self.settings = load_settings()
        self.stop_evt = threading.Event()
        self.worker = None

        self.rows: List[List[Any]] = []
        self.iid_to_row_index: Dict[str,int] = {}
        self.imei_to_iid: Dict[str,str] = {}
        self.iid_to_ids: Dict[str, Dict[str,str]] = {}   # indir butonu için

        self.seen_in_pairs  = set()
        self.seen_out_pairs = set()
        self.imei_first_doc_in: Dict[str,str] = {}
        self.imei_first_doc_out: Dict[str,str] = {}

        self.force_imeis_order: List[str] = []
        self.force_imeis_set: set = set()

        # IMEI bazında KDV ve ipuçları
        self.imei_kdv_in: Dict[str, Set[int]]  = {}
        self.imei_kdv_out: Dict[str, Set[int]] = {}
        self.imei_flags: Dict[str, Dict[str,bool]] = {}  # {"ref":bool,"2el":bool,"gp":bool}

        # Filtreler
        self.docnos_filter: Set[str] = set()
        self.allowed_sources = {"hepsi","sadece_imei","sadece_doc","sadece_url","sadece_excel"}
        self._sort_state: Dict[str, bool] = {}  # column -> asc/desc
        self._all_iids: List[str] = []  # filtre için

        self._build_ui()

    # ---------- UI ----------
    def _build_ui(self):
        # Kimlik
        top = ttk.LabelFrame(self, text="Kimlik & Klasör"); top.pack(fill="x", padx=12, pady=6)
        ttk.Label(top, text="API Token:").grid(row=0, column=0, sticky="e")
        self.tk_token = tk.StringVar(value=self.settings.get("api_token",""))
        ttk.Entry(top, textvariable=self.tk_token, show="•", width=60).grid(row=0, column=1, columnspan=3, sticky="we", padx=6)
        ttk.Button(top, text="Kaydet", command=self._save_settings).grid(row=0, column=4, padx=6)
        ttk.Label(top, text="İndirme Klasörü:").grid(row=0, column=5, sticky="e")
        self.tk_dir = tk.StringVar(value=self.settings.get("download_dir", os.getcwd()))
        ttk.Entry(top, textvariable=self.tk_dir, width=40).grid(row=0, column=6, sticky="we")
        ttk.Button(top, text="Seç", command=self._pick_dir).grid(row=0, column=7, padx=6)
        for c in range(8): top.columnconfigure(c, weight=1)

        # Filtreler (Ünvan/TCKN/VKN + Tarih)
        flt = ttk.LabelFrame(self, text="Filtreler"); flt.pack(fill="x", padx=12, pady=6)
        for c in range(16): flt.columnconfigure(c, weight=1)
        ttk.Label(flt, text="Tedarikçi Ünvan içerir:").grid(row=0, column=0, sticky="e")
        self.tk_unvan = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_unvan, width=26).grid(row=0, column=1, sticky="w")
        ttk.Label(flt, text="TCKN:").grid(row=0, column=2, sticky="e")
        self.tk_tckn = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_tckn, width=14).grid(row=0, column=3, sticky="w")
        ttk.Label(flt, text="VKN:").grid(row=0, column=4, sticky="e")
        self.tk_vkn = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_vkn, width=14).grid(row=0, column=5, sticky="w")

        ttk.Separator(flt, orient="vertical").grid(row=0, column=6, rowspan=2, sticky="ns", padx=8)
        ttk.Label(flt, text="Alıcı Ünvan içerir:").grid(row=0, column=7, sticky="e")
        self.tk_alici_unvan = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_alici_unvan, width=26).grid(row=0, column=8, sticky="w")
        ttk.Label(flt, text="Alıcı TCKN:").grid(row=0, column=9, sticky="e")
        self.tk_alici_tckn = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_alici_tckn, width=14).grid(row=0, column=10, sticky="w")
        ttk.Label(flt, text="Alıcı VKN:").grid(row=0, column=11, sticky="e")
        self.tk_alici_vkn = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_alici_vkn, width=14).grid(row=0, column=12, sticky="w")

        self.tk_use_date = tk.BooleanVar(value=True)
        ttk.Checkbutton(flt, text="Tarih filtresi", variable=self.tk_use_date).grid(row=0, column=13, sticky="w")
        ttk.Label(flt, text="Başlangıç (YYYY-MM-DD):").grid(row=0, column=14, sticky="e")
        self.tk_start = tk.StringVar(value=DEFAULT_DATE_START); ttk.Entry(flt, textvariable=self.tk_start, width=12).grid(row=0, column=15, sticky="w")
        ttk.Label(flt, text="Bitiş:").grid(row=1, column=14, sticky="e")
        self.tk_end = tk.StringVar(value=_today_str()); ttk.Entry(flt, textvariable=self.tk_end, width=12).grid(row=1, column=15, sticky="w")

        # Arama opsiyonları (yeni)
        opts = ttk.LabelFrame(self, text="Arama Opsiyonları"); opts.pack(fill="x", padx=12, pady=6)
        row1 = ttk.Frame(opts); row1.pack(fill="x", padx=8, pady=4)
        self.tk_source_mode = tk.StringVar(value="hepsi")
        for txt, val in [
            ("Hepsi", "hepsi"),
            ("Sadece IMEI listesi", "sadece_imei"),
            ("Sadece Fatura No listesi", "sadece_doc"),
            ("Sadece URL listesi", "sadece_url"),
            ("Sadece Excel (GP)", "sadece_excel"),
        ]:
            ttk.Radiobutton(row1, text=txt, value=val, variable=self.tk_source_mode).pack(side="left", padx=6)

        row2 = ttk.Frame(opts); row2.pack(fill="x", padx=8, pady=4)
        self.tk_include_outside = tk.BooleanVar(value=False)
        self.tk_get_pdf = tk.BooleanVar(value=False)
        self.tk_get_xml = tk.BooleanVar(value=True)
        self.tk_auto_xlsx = tk.BooleanVar(value=True)
        ttk.Checkbutton(row2, text="Listeler dışı da tara?", variable=self.tk_include_outside).pack(side="left", padx=6)
        ttk.Checkbutton(row2, text="PDF indir", variable=self.tk_get_pdf).pack(side="left", padx=6)
        ttk.Checkbutton(row2, text="XML indir", variable=self.tk_get_xml).pack(side="left", padx=6)
        ttk.Checkbutton(row2, text="Bittiğinde otomatik Excel yaz", variable=self.tk_auto_xlsx).pack(side="left", padx=6)

        # IMEI Listesi paneli
        imeif = ttk.LabelFrame(self, text="IMEI Listesi – isteğe bağlı"); imeif.pack(fill="x", padx=12, pady=6)
        r = ttk.Frame(imeif); r.pack(fill="x", padx=8, pady=4)
        self.tk_imei_count = tk.StringVar(value="Seçili IMEI: 0")
        ttk.Button(r, text="TXT/CSV'den Yükle", command=self._load_imeis_from_txtcsv).pack(side="left", padx=4)
        ttk.Button(r, text="Excel'den Yükle", command=self._load_imeis_from_excel).pack(side="left", padx=4)
        ttk.Button(r, text="Panodan Yapıştır", command=self._paste_imeis_clip).pack(side="left", padx=4)
        ttk.Button(r, text="Temizle", command=self._clear_imeis).pack(side="left", padx=4)
        ttk.Label(r, textvariable=self.tk_imei_count).pack(side="left", padx=10)

        # Fatura No Listesi paneli (EAR/EFR)
        docf = ttk.LabelFrame(self, text="Fatura No Listesi (EAR/EFR) – isteğe bağlı"); docf.pack(fill="x", padx=12, pady=6)
        row = ttk.Frame(docf); row.pack(fill="x", padx=8, pady=4)
        self.tk_docnos_count = tk.StringVar(value="Seçili fatura: 0")
        ttk.Button(row, text="TXT'den Yükle", command=self._load_docnos_from_file).pack(side="left", padx=4)
        ttk.Button(row, text="Panodan Yapıştır", command=self._paste_docnos_clip).pack(side="left", padx=4)
        ttk.Button(row, text="Temizle", command=self._clear_docnos).pack(side="left", padx=4)
        ttk.Label(row, textvariable=self.tk_docnos_count).pack(side="left", padx=10)

        # Harici IMEI kaynakları (URL/Excel GP)
        ext = ttk.LabelFrame(self, text="Gider Pusulası Kaynakları (her satıra bir URL)"); ext.pack(fill="x", padx=12, pady=6)
        self.tk_gp_urls = tk.Text(ext, height=3)
        self.tk_gp_urls.pack(fill="x", padx=6, pady=6)
        self.tk_gp_urls.insert("1.0",
            "https://docs.google.com/spreadsheets/d/e/2PACX-1vSDMPeXeKs0HSD38CJst-_1AevO_YuZYtQa7jg-ra0OWQmxi-6qqXGEbqDO_I8ToQ/pub?output=xlsx\n"
        )
        btns = ttk.Frame(ext); btns.pack(fill="x")
        ttk.Button(btns, text="URL’leri Yükle", command=self._load_gp_from_urls).pack(side="left", padx=6)
        ttk.Button(btns, text="Excel’den Yükle (dosya)", command=self._load_gp_from_file).pack(side="left", padx=6)

        # Aksiyonlar
        act = ttk.Frame(self); act.pack(fill="x", padx=12, pady=4)
        ttk.Button(act, text="Tara (XML içinden)", command=self._start_scan).pack(side="left", padx=4)
        ttk.Button(act, text="Seçili → PDF indir", command=lambda: self._download_selected(mode="pdf")).pack(side="left", padx=4)
        ttk.Button(act, text="Seçili → XML indir", command=lambda: self._download_selected(mode="xml")).pack(side="left", padx=4)
        ttk.Button(act, text="Seçili → PDF+XML", command=lambda: self._download_selected(mode="both")).pack(side="left", padx=4)
        ttk.Button(act, text="Excel'e Aktar", command=self._export_excel).pack(side="left", padx=4)
        self.btn_stop = ttk.Button(act, text="Durdur", command=self._stop_now, state="disabled"); self.btn_stop.pack(side="right", padx=4)

        # PanedWindow: Üst → Tablo; Alt → Log
        pan = ttk.Panedwindow(self, orient="vertical"); pan.pack(fill="both", expand=True, padx=12, pady=(0,6))

        # Üst panel: filtre çubuğu + tablo
        top_panel = ttk.Frame(pan)
        # Filtre çubuğu
        bar = ttk.Frame(top_panel); bar.pack(fill="x", pady=(6,4))
        ttk.Label(bar, text="Metin Filtresi:").pack(side="left", padx=(0,6))
        self.tk_filter = tk.StringVar()
        ent = ttk.Entry(bar, textvariable=self.tk_filter, width=40); ent.pack(side="left")
        self.tk_filter.trace_add("write", lambda *_: self._apply_filter())
        ttk.Button(bar, text="Temizle", command=lambda: (self.tk_filter.set(""), self._apply_filter())).pack(side="left", padx=6)
        self.tk_only_incomplete = tk.BooleanVar(value=False)
        ttk.Checkbutton(bar, text="Eksik bilgili satırları göster", variable=self.tk_only_incomplete, command=self._apply_filter).pack(side="left", padx=10)

        # Tablo
        self.tree = ttk.Treeview(top_panel, columns=HEADERS, show="headings", height=18, selectmode="extended")
        for c in HEADERS:
            self.tree.heading(c, text=c, command=lambda col=c: self._sort_by(col))
            w = 150
            if c in ("Marka","MODEL","GEREKÇE"): w = 320
            if c in ("Borç Tutar","SATIŞ BEDELİ","KDV TUTARI"): w = 120
            if c in ("Sütun1",): w = 320
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=0, pady=(0,6))
        ysb = ttk.Scrollbar(self.tree, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=ysb.set); ysb.place(in_=self.tree, relx=1.0, rely=0, relheight=1.0, x=-16)
        # Kopyalama ve indirme kısayolları
        self.tree.bind("<Control-c>", self._copy_selection)
        self.tree.bind("<Button-3>", self._popup_menu)
        self.tree.bind("<Double-1>", lambda e: self._download_selected(mode="pdf"))

        pan.add(top_panel, weight=5)

        # Alt panel: Log
        logf = ttk.Frame(pan)
        lab = ttk.Label(logf, text="Log", anchor="w"); lab.pack(fill="x")
        self.log = scrolledtext.ScrolledText(logf, height=10); self.log.pack(fill="both", expand=True, padx=0, pady=(0,6))
        self.logger = setup_logger(self.log)
        pan.add(logf, weight=2)

    # ---------- helpers ----------
    def _log(self, msg: str): self.logger.info(msg)
    def _save_settings(self):
        s = self.settings
        s["api_token"] = self.tk_token.get().strip()
        s["download_dir"] = self.tk_dir.get().strip() or os.getcwd()
        s["out_name"] = s.get("out_name", DEFAULTS["out_name"])
        s["timeout_connect"] = int(DEFAULTS["timeout_connect"])
        s["timeout_read"]    = int(DEFAULTS["timeout_read"])
        s["retries"]         = int(DEFAULTS["retries"])
        s["backoff"]         = float(DEFAULTS["backoff"])
        save_settings(s); messagebox.showinfo("Bilgi", "Ayarlar kaydedildi.")
    def _pick_dir(self):
        d = filedialog.askdirectory(title="İndirme klasörü")
        if d: self.tk_dir.set(d)
    def _stop_now(self):
        self.stop_evt.set(); self._log("❌ Durdurma istendi. Mevcut sayfa bitince durur."); self.btn_stop.config(state="disabled")

    # ---------- IMEI Listesi yönetimi ----------
    def _set_imeis(self, arr: List[str], append: bool = True):
        if not append:
            self.force_imeis_order = []
            self.force_imeis_set = set()
        for im in arr:
            if re.fullmatch(r"\d{15}", im) and _luhn_ok_imei(im):
                if im not in self.force_imeis_set:
                    self.force_imeis_order.append(im)
                    self.force_imeis_set.add(im)
        self.tk_imei_count.set(f"Seçili IMEI: {len(self.force_imeis_set)}")
        self._log(f"[IMEI] Listede {len(self.force_imeis_set)} IMEI var.")

    def _load_imeis_from_txtcsv(self):
        p = filedialog.askopenfilename(title="IMEI listesi (TXT/CSV)", filetypes=[("Metin/CSV","*.txt *.csv"),("Tümü","*.*")])
        if not p: return
        arr: List[str] = []
        try:
            ext = os.path.splitext(p)[1].lower()
            if ext == ".csv":
                with open(p, "r", encoding="utf-8") as f:
                    for row in csv.reader(f):
                        if row:
                            v = (row[0] or "").strip()
                            if re.fullmatch(r"\d{15}", v) and _luhn_ok_imei(v): arr.append(v)
            else:
                with open(p, "r", encoding="utf-8") as f:
                    for line in f:
                        v = line.strip()
                        if re.fullmatch(r"\d{15}", v) and _luhn_ok_imei(v): arr.append(v)
        except Exception as e:
            messagebox.showerror("Hata", f"IMEI listesi okunamadı: {e}"); return
        self._set_imeis(arr, append=False)

    def _load_imeis_from_excel(self):
        p = filedialog.askopenfilename(title="IMEI listesi (Excel)", filetypes=[("Excel","*.xlsx *.xls")])
        if not p: return
        arr: List[str] = []
        try:
            wb = load_workbook(p, data_only=True); ws = wb.active
            for r in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                v = str(r[0]).strip() if r and r[0] is not None else ""
                if re.fullmatch(r"\d{15}", v) and _luhn_ok_imei(v): arr.append(v)
        except Exception as e:
            messagebox.showerror("Hata", f"Excel okunamadı: {e}"); return
        self._set_imeis(arr, append=False)

    def _paste_imeis_clip(self):
        try:
            txt = self.clipboard_get()
        except Exception:
            messagebox.showerror("Hata", "Panoda veri yok."); return
        ims = [m.group(0) for m in IMEI_RE_STRICT.finditer(txt) if _luhn_ok_imei(m.group(0))]
        if not ims:
            messagebox.showinfo("Bilgi", "Panodaki metinde IMEI bulunamadı.")
        self._set_imeis(ims, append=False)

    def _clear_imeis(self):
        self.force_imeis_order = []; self.force_imeis_set = set()
        self.tk_imei_count.set("Seçili IMEI: 0")
        self._log("[IMEI] Liste temizlendi.")

    # ---------- Fatura no filtresi (EAR/EFR) ----------
    def _extract_docnos(self, text: str) -> List[str]:
        if not text: return []
        seen = set(); out = []
        for m in DOCNO_RE.findall(text.upper()):
            if m not in seen:
                seen.add(m); out.append(m)
        return out
    def _set_docnos(self, arr: List[str]):
        self.docnos_filter = set(arr or [])
        self.tk_docnos_count.set(f"Seçili fatura: {len(self.docnos_filter)}")
        if arr:
            self._log(f"[FILTRE] {len(arr)} EAR/EFR fatura yüklendi.")
        else:
            self._log("[FILTRE] Fatura listesi temizlendi.")
    def _load_docnos_from_file(self):
        p = filedialog.askopenfilename(title="Fatura Listesi (.txt)", filetypes=[("Metin","*.txt"),("Tümü","*.*")])
        if not p: return
        try:
            with open(p,"r",encoding="utf-8") as f:
                arr = self._extract_docnos(f.read())
            self._set_docnos(arr)
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya okunamadı: {e}")
    def _paste_docnos_clip(self):
        try:
            txt = self.clipboard_get()
        except Exception:
            messagebox.showerror("Hata", "Panoda veri yok."); return
        arr = self._extract_docnos(txt)
        if not arr:
            messagebox.showinfo("Bilgi", "Panodaki metinde EAR/EFR bulunamadı.")
        self._set_docnos(arr)
    def _clear_docnos(self):
        self._set_docnos([])

    # ---------- GP yükleme (URL/Dosya) [aynı mantık] ----------
    def _merge_gp_ready_rows(self, rows_ready: List[List[Any]]):
        added = merged = 0
        for ready in rows_ready:
            ready = ensure_len(ready)
            rd = {h: (ready[i] if i < len(ready) else "") for i, h in enumerate(HEADERS)}
            im = rd["imei"]
            if not im: continue
            # ipuçları
            self._mark_flags(im, ref=bool(KEY_REF.search(nup(rd.get("MODEL","")))),
                                 is2=bool(KEY_2EL.search(nup(rd.get("MODEL","")))),
                                 gp=True)
            iid = self.imei_to_iid.get(im)
            if iid:
                cur = ensure_len(list(self.tree.item(iid, "values")))
                for i, h in enumerate(HEADERS[:20]):
                    if not cur[i] and rd.get(h): cur[i] = rd[h]
                cur[1] = "XML+GP" if (cur[1] and cur[1] != "GP") else (cur[1] or "GP")
                self.tree.item(iid, values=cur)
                idxr = self.iid_to_row_index.get(iid)
                if idxr is not None: self.rows[idxr] = ensure_len(cur)
                merged += 1
            else:
                row = [rd.get(h, "") for h in HEADERS]
                row = ensure_len(row)
                iid = self.tree.insert("", "end", values=row)
                self.iid_to_row_index[iid] = len(self.rows)
                self.rows.append(row)
                self.imei_to_iid[im] = iid
                added += 1
            self.force_imeis_order.append(im); self.force_imeis_set.add(im)
            self._update_classification_for(im)
        self._log(f"[GP] Yüklendi (şablon uyumlu): yeni={added}, birleştirilen={merged}")

    def _load_gp_from_urls(self):
        urls = [u.strip() for u in self.tk_gp_urls.get("1.0","end").splitlines() if u.strip()]
        if not urls:
            messagebox.showinfo("Bilgi", "Önce en az bir URL girin."); return
        for url in urls:
            try:
                self._log(f"[GP] URL indiriliyor: {url}")
                resp = http_get(url, token=None, typ="bin", log=self._log, stop_evt=self.stop_evt)
                if resp is None:
                    self._log(f"[GP] ❌ URL yüklenemedi (timeout/hata): {url}")
                    continue
                wb = load_workbook(io.BytesIO(resp.content), data_only=True)
                rows_ready = parse_gp_template_workbook(wb, self._log)
                if rows_ready:
                    self._merge_gp_ready_rows(rows_ready)
                else:
                    items = parse_gp_workbook(wb, self._log)
                    self._merge_gp_items(items)
            except Exception as e:
                self._log(f"[GP] ❌ URL yüklenemedi: {e}")
        self._log(f"[GP] Tamamlandı.")

    def _load_gp_from_file(self):
        p = filedialog.askopenfilename(title="Gider Pusulası Excel seç", filetypes=[("Excel","*.xlsx *.xls")])
        if not p: return
        try:
            wb = load_workbook(p, data_only=True)
            rows_ready = parse_gp_template_workbook(wb, self._log)
            if rows_ready:
                self._merge_gp_ready_rows(rows_ready); return
            items = parse_gp_workbook(wb, self._log)
            self._merge_gp_items(items)
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya yüklenemedi: {e}")

    def _merge_gp_items(self, items: List[Dict[str,Any]]):
        if not items:
            self._log("[GP] Kayıt bulunamadı."); return
        added = merged = 0
        for it in items:
            im = it["imei"]
            self.force_imeis_order.append(im); self.force_imeis_set.add(im)
            # İpuçları
            txt = nup(it.get("aciklama",""))
            self._mark_flags(im, ref=bool(KEY_REF.search(txt)), is2=bool(KEY_2EL.search(txt)), gp=True)

            iid = self.imei_to_iid.get(im)
            brand = brand_from_text(it.get("aciklama",""))
            model = it.get("aciklama","")
            borc  = it.get("bedel","")
            sube  = it.get("sube","")
            info  = "; ".join([v for v in [("Şube: "+sube) if sube else "", it.get("aciklama","")] if v])

            if iid:
                vals = ensure_len(list(self.tree.item(iid, "values")))
                if not vals[7] and borc:  vals[7] = borc
                if vals[17]: vals[17] += (" | " + info) if info else ""
                else: vals[17] = info
                if vals[18] and "Gider Pusulası" not in vals[18]:
                    vals[18] = vals[18] + " + Gider Pusulası"
                elif not vals[18]:
                    vals[18] = "Gider Pusulası"
                if not vals[8] or vals[8]=="Bilinmeyen": vals[8] = brand
                if not vals[9]: vals[9] = model
                if not vals[3]: vals[3] = "GİDER PUSULASI"
                if not vals[19]: vals[19] = "Satılabilir"
                vals[1] = "XML+GP" if (vals[1] and vals[1] != "GP") else (vals[1] or "GP")
                self.tree.item(iid, values=vals)
                idx = self.iid_to_row_index.get(iid)
                if idx is not None: self.rows[idx] = ensure_len(vals)
                merged += 1
            else:
                row = [
                    im, "GP", "", "GİDER PUSULASI", it.get("tarih",""), "GP", it.get("ad",""),
                    borc, brand, model,
                    "", "", "", "", "", "", "",
                    info, "Gider Pusulası", "Satılabilir",
                    "", "", "", ""
                ]
                row = ensure_len(row)
                iid = self.tree.insert("", "end", values=row)
                self.iid_to_row_index[iid] = len(self.rows)
                self.rows.append(row)
                self.imei_to_iid.setdefault(im, iid)
                added += 1

            self._update_classification_for(im)
        self._log(f"[GP] Yüklendi: {len(items)} kayıt | yeni={added}, birleştirilen={merged}")

    # ---------- İpuçları & KDV setleri ----------
    def _mark_flags(self, imei: str, *, ref: Optional[bool]=None, is2: Optional[bool]=None, gp: Optional[bool]=None):
        f = self.imei_flags.get(imei, {"ref":False,"2el":False,"gp":False})
        if ref is not None: f["ref"] = f["ref"] or ref
        if is2 is not None: f["2el"] = f["2el"] or is2
        if gp  is not None: f["gp"]  = f["gp"]  or gp
        self.imei_flags[imei] = f

    def _add_kdv_in(self, imei: str, kdv: Optional[float]):
        if kdv is None: return
        s = self.imei_kdv_in.get(imei, set()); s.add(int(round(kdv))); self.imei_kdv_in[imei] = s

    def _add_kdv_out(self, imei: str, kdv: Optional[float]):
        if kdv is None: return
        s = self.imei_kdv_out.get(imei, set()); s.add(int(round(kdv))); self.imei_kdv_out[imei] = s

    def _find_kdv_for_imei(self, P: Parsed, imei: str) -> Optional[float]:
        for L in P.lines:
            if imei and imei in L["blob"]:
                return L.get("kdv", None) or P.inv_kdv
        return P.inv_kdv

    def _stringify_kdvset(self, s: Set[int]) -> str:
        if not s: return ""
        return ",".join([str(x) for x in sorted(s)])

    def _update_kdv_cols(self, imei: str):
        iid = self.imei_to_iid.get(imei)
        if not iid: return
        vals = ensure_len(list(self.tree.item(iid, "values")))
        vals[20] = self._stringify_kdvset(self.imei_kdv_in.get(imei, set()))
        vals[21] = self._stringify_kdvset(self.imei_kdv_out.get(imei, set()))
        self.tree.item(iid, values=vals)
        idx = self.iid_to_row_index.get(iid)
        if idx is not None: self.rows[idx] = ensure_len(vals)

    def _update_classification_for(self, imei: str):
        iid = self.imei_to_iid.get(imei)
        if not iid: return
        vals = ensure_len(list(self.tree.item(iid, "values")))

        k_in  = self.imei_kdv_in.get(imei, set())
        k_out = self.imei_kdv_out.get(imei, set())
        f = self.imei_flags.get(imei, {"ref":False,"2el":False,"gp":False})
        reasons = []
        klass = ""

        has_satis_1 = (1 in k_out)
        has_alis_1  = (1 in k_in)
        has_alis_20 = (20 in k_in)
        has_satis_20= (20 in k_out)

        # ---- YENİLENMİŞ ----
        if has_satis_1:
            klass = "YENİLENMİŞ"
            reasons.append("SATIŞ KDV=1")
            if has_alis_1:
                reasons.append("ALIŞ 1 → SATIŞ 1")
            if has_alis_20:
                reasons.append("ALIŞ 20 → SATIŞ 1")
            if f.get("gp"): reasons.append("Gider pusulası + SATIŞ 1")
        if not klass and f.get("ref"):
            klass = "YENİLENMİŞ"
            reasons.append("Metin ipucu: YENİLENMİŞ/REFURB")

        # ---- 2.EL ----
        if not klass and f.get("2el"):
            klass = "2.EL"
            reasons.append("Metin ipucu: 2.EL")
            if has_alis_20 and has_satis_20:
                reasons.append("ALIŞ 20 → SATIŞ 20")

        vals[20] = self._stringify_kdvset(k_in)
        vals[21] = self._stringify_kdvset(k_out)
        vals[22] = klass
        vals[23] = "; ".join(reasons)
        self.tree.item(iid, values=vals)
        idx = self.iid_to_row_index.get(iid)
        if idx is not None: self.rows[idx] = ensure_len(vals)

    # ---------- Akış ----------
    def _start_scan(self):
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("Bilgi", "Devam eden iş var. Önce durdurun."); return
        if not self.tk_token.get().strip():
            messagebox.showwarning("Uyarı", "Önce API token girin."); return

        # Ağ ayarlarını uygula
        TIMEOUT_CONNECT = int(self.settings.get("timeout_connect", DEFAULTS["timeout_connect"]))
        TIMEOUT_READ    = int(self.settings.get("timeout_read", DEFAULTS["timeout_read"]))
        RETRIES         = int(self.settings.get("retries", DEFAULTS["retries"]))
        BACKOFF         = float(self.settings.get("backoff", DEFAULTS["backoff"]))
        SESSION = make_session(RETRIES, BACKOFF)

        self.log.delete("1.0", tk.END); self.stop_evt.clear(); self.btn_stop.config(state="normal")

        # Kaynak modu ve kapsama
        self._scan_mode = self.tk_source_mode.get()
        self._include_outside = self.tk_include_outside.get()

        # Boş tabloyu koru (mevcut veriler kalsın) – yeni tarama sadece ekler/birleştirir
        self.worker = threading.Thread(target=self._scan_flow, daemon=True); self.worker.start()

    def _append_or_merge_purchase(self, P: Parsed, inv_id: str, doc_no: str, im: str, borc_tutar: str, model: str, brand: str):
        key = (doc_no, im)
        if key in self.seen_in_pairs:
            self._log(f"    ↳ IMEI {im} ⚠ zaten eklendi (alış mükerrer): {doc_no}")
            iid = self.imei_to_iid.get(im)
            if iid:
                vals = ensure_len(list(self.tree.item(iid, "values")))
                vals[17] = (vals[17] + " | " if vals[17] else "") + f"Alışta mükerrer: {doc_no}"
                self.tree.item(iid, values=vals)
                idxr = self.iid_to_row_index.get(iid)
                if idxr is not None: self.rows[idxr] = ensure_len(vals)
            return
        self.seen_in_pairs.add(key)

        kdv = self._find_kdv_for_imei(P, im)
        self._add_kdv_in(im, kdv)
        self._mark_flags(im, ref=bool(KEY_REF.search(P.text_upper)), is2=bool(KEY_2EL.search(P.text_upper)), gp=False)

        iid = self.imei_to_iid.get(im)
        if iid:
            vals = ensure_len(list(self.tree.item(iid, "values")))
            vals[1] = "XML+GP" if (vals[1] and vals[1] != "GP") else (vals[1] or "XML")
            if not vals[2]: vals[2] = P.supplier_id
            vals[3] = "FATURA" if vals[3] != "GİDER PUSULASI" else vals[3]
            if not vals[4]: vals[4] = P.issue_date
            if not vals[5]: vals[5] = P.invoice_no or doc_no
            if not vals[6]: vals[6] = P.supplier_name
            if not vals[7] and borc_tutar: vals[7] = borc_tutar
            if not vals[8] or vals[8]=="Bilinmeyen": vals[8] = brand
            if not vals[9]: vals[9] = model
            vals[18] = "Fatura + Gider Pusulası" if "Gider Pusulası" in (vals[18] or "") else (vals[18] or "Fatura")
            self.tree.item(iid, values=vals)
            idxr = self.iid_to_row_index.get(iid)
            if idxr is not None: self.rows[idxr] = ensure_len(vals)
        else:
            row = [
                im, "XML", P.supplier_id, "FATURA", P.issue_date, P.invoice_no or doc_no, P.supplier_name,
                borc_tutar, brand, model,
                "", "", "", "", "", "", "",
                "", "Fatura", "Satılabilir",
                "", "", "", ""
            ]
            row = ensure_len(row)
            iid = self.tree.insert("", "end", values=row)
            self.iid_to_row_index[iid] = len(self.rows)
            self.rows.append(row)
            self.imei_to_iid.setdefault(im, iid)

        info = self.iid_to_ids.get(iid, {})
        info.update({"in_id": inv_id, "in_doc": (P.invoice_no or doc_no)})
        self.iid_to_ids[iid] = info

        self._update_kdv_cols(im)
        self._update_classification_for(im)

    def _append_or_merge_sale(self, P: Parsed, inv_id: str, doc_no: str, im: str, kind: str):
        key = (doc_no, im)
        if key in self.seen_out_pairs:
            self._log(f"    ↳ IMEI {im} ⚠ zaten eklendi (satış mükerrer): {doc_no}")
            iid = self.imei_to_iid.get(im)
            if iid:
                vals = ensure_len(list(self.tree.item(iid, "values")))
                vals[17] = (vals[17] + " | " if vals[17] else "") + f"Satışta mükerrer: {doc_no}"
                self.tree.item(iid, values=vals)
                idxr = self.iid_to_row_index.get(iid)
                if idxr is not None: self.rows[idxr] = ensure_len(vals)
            return
        self.seen_out_pairs.add(key)

        kdv = self._find_kdv_for_imei(P, im)
        self._add_kdv_out(im, kdv)
        self._mark_flags(im, ref=bool(KEY_REF.search(P.text_upper)), is2=bool(KEY_2EL.search(P.text_upper)), gp=False)

        iid = self.imei_to_iid.get(im)
        if not iid:
            row = [
                im,"XML_ONLY","", "", "", "", "",
                "", brand_from_text(" ".join(P.items)), P.model or (P.items[0] if P.items else ""),
                P.issue_date, P.buyer_name, P.payable, P.tax_total, P.invoice_no or doc_no,
                P.buyer_id_type, P.buyer_id,
                "", "", "Alış kaydı eksik",
                "", "", "", ""
            ]
            row = ensure_len(row)
            iid = self.tree.insert("", "end", values=row)
            self.iid_to_row_index[iid] = len(self.rows)
            self.rows.append(row)
            self.imei_to_iid.setdefault(im, iid)

        vals = ensure_len(list(self.tree.item(iid, "values")))
        if not vals[14]:
            vals[10] = P.issue_date
            vals[11] = P.buyer_name
            vals[12] = P.payable
            vals[13] = P.tax_total
            vals[14] = P.invoice_no or doc_no
            vals[15] = P.buyer_id_type
            vals[16] = P.buyer_id
            vals[19] = "Satılmış"
        else:
            vals[17] = (vals[17] + " | " if vals[17] else "") + f"⚠ IMEI başka satışta da var: {vals[14]}"
        self.tree.item(iid, values=vals)
        idxr = self.iid_to_row_index.get(iid)
        if idxr is not None: self.rows[idxr] = ensure_len(vals)
        info = self.iid_to_ids.get(iid, {})
        info.update({"out_id": inv_id, "out_doc": (P.invoice_no or doc_no), "out_kind": kind})
        self.iid_to_ids[iid] = info

        first_s = self.imei_first_doc_out.get(im)
        if first_s and first_s != doc_no:
            vals = ensure_len(list(self.tree.item(iid, "values")))
            vals[17] = (vals[17] + " | " if vals[17] else "") + f"⚠ IMEI başka satışta da var: {first_s}"
            self.tree.item(iid, values=vals)
            if idxr is not None: self.rows[idxr] = ensure_len(vals)
        else:
            self.imei_first_doc_out.setdefault(im, doc_no)

        self._update_kdv_cols(im)
        self._update_classification_for(im)

    def _scan_flow(self):
        try:
            token = self.tk_token.get().strip()
            start = self.tk_start.get().strip() if self.tk_use_date.get() else ""
            end   = self.tk_end.get().strip() if self.tk_use_date.get() else ""

            mode = self._scan_mode
            include_outside = self._include_outside

            self._log(f"🔷 Kaynak modu: {mode} | Listeler dışı da tara? {'Evet' if include_outside else 'Hayır'}")

            # SATIŞ ÖNCE? Yalnızca 'sadece_doc' modunda satıştan başla, diğerlerinde IMEI odaklı
            if mode in ("hepsi","sadece_doc"):
                # e-Arşiv (satış) — XML içinden filtre
                out_arch = list_both_archived(EARCH_OUT_LIST, token, start, end, log=self._log, stop_evt=self.stop_evt, section_name="SATIŞ-eArşiv")
                self._log(f"[SATIŞ/e-Arşiv] Tekil meta: {len(out_arch)}")
                for idx, meta in enumerate(out_arch, 1):
                    if self.stop_evt.is_set(): break
                    inv_id = str(meta.get("id") or ""); doc_no_meta = str(meta.get("documentNumber") or inv_id)
                    xmlb = fetch_xml_by(EARCH_OUT_DOC, token, inv_id)
                    if not xmlb: 
                        self._log(f"[E-ARŞİV] {idx}/{len(out_arch)} {doc_no_meta}: XML indirilemedi."); 
                        continue
                    P = parse_invoice_xml(xmlb)
                    # Fatura No listesi varsa, filtreyi XML içinden (P.invoice_no) uygula
                    if self.docnos_filter and (P.invoice_no or "").upper() not in self.docnos_filter:
                        continue
                    # Alıcı filtreleri (XML'den)
                    f_unvan = nlow(self.tk_alici_unvan.get()); f_tckn = self.tk_alici_tckn.get().strip(); f_vkn = self.tk_alici_vkn.get().strip()
                    if f_unvan and (f_unvan not in nlow(P.buyer_name)): continue
                    if f_tckn and P.buyer_id != f_tckn: continue
                    if f_vkn  and P.buyer_id != f_vkn:  continue
                    if not P.imeis: 
                        self._log(f"[E-ARŞİV] {(P.invoice_no or doc_no_meta)}: IMEI yok."); 
                        continue
                    # IMEI kapsamı: Eğer include_outside False ise ve IMEI listesi varsa, sadece listedekiler
                    for im in P.imeis:
                        if not include_outside and self.force_imeis_set and im not in self.force_imeis_set:
                            continue
                        self.force_imeis_order.append(im); self.force_imeis_set.add(im)
                        self._append_or_merge_sale(P, inv_id, P.invoice_no or doc_no_meta, im, kind="E-ARŞİV")

                # E-Fatura Giden (satış)
                out_einv = list_both_archived(EINV_OUT_LIST, token, start, end, log=self._log, stop_evt=self.stop_evt, section_name="SATIŞ-Giden")
                self._log(f"[SATIŞ/E-Fatura Giden] Tekil meta: {len(out_einv)}")
                for idx, meta in enumerate(out_einv, 1):
                    if self.stop_evt.is_set(): break
                    inv_id = str(meta.get("id") or ""); doc_no_meta = str(meta.get("documentNumber") or inv_id)
                    xmlb = fetch_xml_by(EINV_OUT_DOC, token, inv_id)
                    if not xmlb: 
                        self._log(f"[E-FATURA] {idx}/{len(out_einv)} {doc_no_meta}: XML indirilemedi."); 
                        continue
                    P = parse_invoice_xml(xmlb)
                    if self.docnos_filter and (P.invoice_no or "").upper() not in self.docnos_filter:
                        continue
                    f_unvan = nlow(self.tk_alici_unvan.get()); f_tckn = self.tk_alici_tckn.get().strip(); f_vkn = self.tk_alici_vkn.get().strip()
                    if f_unvan and (f_unvan not in nlow(P.buyer_name)): continue
                    if f_tckn and P.buyer_id != f_tckn: continue
                    if f_vkn  and P.buyer_id != f_vkn:  continue
                    if not P.imeis:
                        self._log(f"[E-FATURA] {(P.invoice_no or doc_no_meta)}: IMEI yok."); 
                        continue
                    for im in P.imeis:
                        if not include_outside and self.force_imeis_set and im not in self.force_imeis_set:
                            continue
                        self.force_imeis_order.append(im); self.force_imeis_set.add(im)
                        self._append_or_merge_sale(P, inv_id, P.invoice_no or doc_no_meta, im, kind="E-FATURA")

            # ALIŞ – incoming
            if mode in ("hepsi","sadece_imei","sadece_url","sadece_excel","sadece_doc"):
                self._log("🔹 [ALIŞ] Tarama başlıyor...")
                incoming = list_both_archived(EINV_IN_LIST, token, start, end, log=self._log, stop_evt=self.stop_evt, section_name="ALIŞ")
                self._log(f"🔹 [ALIŞ] Toplam tekil fatura: {len(incoming)}")

                found_imeis = set()
                refurb_rows: List[List[Any]] = []
                renewal_rows: List[List[Any]] = []

                # IMEI kapsamı: listedekiler (IMEI listesi + GP'den gelenler + satıştan toplananlar)
                allowed_imeis = set(self.force_imeis_set)
                only_from_lists = (not include_outside) and bool(allowed_imeis or self.docnos_filter or self.tk_gp_urls.get("1.0","end").strip())

                for idx, meta in enumerate(incoming, 1):
                    if self.stop_evt.is_set(): break
                    inv_id = str(meta.get("id") or ""); doc_no = str(meta.get("documentNumber") or inv_id)
                    xmlb = fetch_xml_by(EINV_IN_DOC, token, inv_id)
                    if not xmlb:
                        self._log(f"[ALIŞ] {idx}/{len(incoming)} {doc_no}: XML indirilemedi."); continue
                    P = parse_invoice_xml(xmlb)

                    # IMEI yoksa ve whitelist tedarikçi ise atla
                    if not P.imeis and is_whitelisted_supplier(P.supplier_name, self.settings):
                        self._log(f"[ALIŞ] {doc_no} | {P.supplier_name} → beyaz liste & IMEI=0: atlandı.")
                        continue

                    # Tedarikçi filtreleri
                    if self.tk_unvan.get().strip() and (nlow(self.tk_unvan.get()) not in nlow(P.supplier_name)): continue
                    if self.tk_tckn.get().strip() and P.supplier_id != self.tk_tckn.get().strip(): continue
                    if self.tk_vkn.get().strip() and P.supplier_id != self.tk_vkn.get().strip(): continue

                    imeis = P.imeis or []
                    # Modlara göre kapsam daraltma
                    if only_from_lists:
                        imeis = [x for x in imeis if (x in allowed_imeis)]
                        if not imeis:
                            continue

                    for im in imeis:
                        unit_price = ""; line_total = ""; model = P.model; brand = P.brand
                        for L in P.lines:
                            if im in L["blob"]:
                                unit_price = L.get("unit_price","") or ""
                                line_total = L.get("line_total","") or ""
                                model = L["blob"]; brand = brand_from_text(L["blob"])
                                break
                        borc_tutar = unit_price or line_total or P.payable
                        self._log(f"    ↳ IMEI {im} → '{borc_tutar}' | Marka={brand} | KDV={self._find_kdv_for_imei(P, im)}")
                        self._append_or_merge_purchase(P, inv_id, P.invoice_no or doc_no, im, borc_tutar, model, brand)
                        found_imeis.add(im)

                    # IMEI yoksa ama yenileme ipuçları
                    if not imeis:
                        tU = P.text_upper
                        if KEY_REF.search(tU):
                            row = ["","XML",P.supplier_id,"FATURA",P.issue_date,P.invoice_no or doc_no,P.supplier_name,
                                   P.payable, brand_from_text(tU), P.model or (P.items[0] if P.items else ""),
                                   "","","","","","","",
                                   "YENİLENMİŞ ürün (IMEI yok)","Fatura","Satılabilir",
                                   "","","",""]
                            refurb_rows.append(ensure_len(row))
                        if "CEP TELEFONU YENİLEME HİZMETİ" in tU:
                            row = ["","XML",P.supplier_id,"FATURA",P.issue_date,P.invoice_no or doc_no,P.supplier_name,
                                   P.payable, brand_from_text(tU), P.model or (P.items[0] if P.items else ""),
                                   "","","","","","","",
                                   "CEP TELEFONU YENİLEME HİZMETİ","Fatura","Satılabilir",
                                   "","","",""]
                            renewal_rows.append(ensure_len(row))

                # Yüklenen IMEI listesinde olup alışta bulunamayanlar
                if self.force_imeis_order:
                    miss = [im for im in self.force_imeis_order if im not in found_imeis]
                    if miss:
                        self._log(f"[ALIŞ] IMEI bulunamadı: {len(miss)} (örnek: {', '.join(miss[:5])})")
                        for im in miss:
                            if im in self.imei_to_iid:
                                iid = self.imei_to_iid[im]
                                vals = ensure_len(list(self.tree.item(iid, "values")))
                                vals[1] = vals[1] or "GP"
                                self.tree.item(iid, values=vals)
                                continue
                            row = [im,"Bulunamadı","","","","","","","","","","","","","","","",
                                   "IMEI listesi/GP var ama alış faturasında bulunamadı","","",
                                   "","","",""]
                            row = ensure_len(row)
                            iid = self.tree.insert("", "end", values=row)
                            self.iid_to_row_index[iid] = len(self.rows)
                            self.rows.append(row)
                            self.imei_to_iid.setdefault(im, iid)

                for r in refurb_rows + renewal_rows:
                    iid = self.tree.insert("", "end", values=ensure_len(r))
                    self.iid_to_row_index[iid] = len(self.rows)
                    self.rows.append(ensure_len(r))

            self._log("✅ Tarama bitti.")
            if self.tk_auto_xlsx.get():
                outp = self.settings.get("out_name", DEFAULTS["out_name"])
                try:
                    write_excel(self._table_rows(), outp); self._log(f"🧾 Excel yazıldı: {outp}")
                except Exception as e:
                    self._log(f"❌ Excel yazılamadı: {e}")

        except Exception as e:
            messagebox.showerror("Hata", str(e))
        finally:
            self.btn_stop.config(state="disabled")

    # ---------- İndirme & Export ----------
    def _download_selected(self, mode: str = "both"):
        sel = self.tree.selection()
        if not sel: 
            messagebox.showinfo("Bilgi", "Listeden en az bir satır seçin."); 
            return
        token = self.tk_token.get().strip()
        if not token: 
            messagebox.showwarning("Uyarı", "Önce token girin."); 
            return
        want_pdf = (mode in ("pdf","both"))
        want_xml = (mode in ("xml","both"))
        save_dir = self.tk_dir.get().strip() or os.getcwd()
        os.makedirs(save_dir, exist_ok=True)

        saved = 0
        for iid in sel:
            ids = self.iid_to_ids.get(iid, {})
            in_id   = ids.get("in_id");  in_doc  = safe_filename(ids.get("in_doc","ALIS"))
            out_id  = ids.get("out_id"); out_doc = safe_filename(ids.get("out_doc","SATIS"))
            out_kind= ids.get("out_kind","E-ARŞİV")

            if in_id:
                if want_xml:
                    xmlb = fetch_xml_by(EINV_IN_DOC, token, in_id)
                    if xmlb:
                        with open(os.path.join(save_dir, f"{in_doc}_ALIS.xml"), "wb") as f: f.write(xmlb); saved += 1
                if want_pdf:
                    pdfb = fetch_pdf_by(EINV_IN_DOC, token, in_id)
                    if pdfb:
                        with open(os.path.join(save_dir, f"{in_doc}_ALIS.pdf"), "wb") as f: f.write(pdfb); saved += 1
            if out_id:
                DOC = EARCH_OUT_DOC if out_kind=="E-ARŞİV" else EINV_OUT_DOC
                if want_xml:
                    xmlb = fetch_xml_by(DOC, token, out_id)
                    if xmlb:
                        with open(os.path.join(save_dir, f"{out_doc}_SATIS.xml"), "wb") as f: f.write(xmlb); saved += 1
                if want_pdf:
                    pdfb = fetch_pdf_by(DOC, token, out_id)
                    if pdfb:
                        with open(os.path.join(save_dir, f"{out_doc}_SATIS.pdf"), "wb") as f: f.write(pdfb); saved += 1

        self._log(f"📦 İndirme tamamlandı. Kaydedilen dosya: {saved}")
        if saved:
            try: os.startfile(save_dir)
            except Exception: pass

    def _table_rows(self) -> List[List[Any]]:
        rows=[]
        for iid in self.tree.get_children(""):
            rows.append(ensure_len(list(self.tree.item(iid, "values"))))
        return rows

    def _export_excel(self):
        rows = self._table_rows()
        if not rows: messagebox.showinfo("Bilgi", "Henüz sonuç yok."); return
        p = filedialog.asksaveasfilename(
            title="Excel kaydet", defaultextension=".xlsx",
            initialfile=os.path.basename(self.settings.get("out_name", DEFAULTS["out_name"])),
            filetypes=[("Excel", "*.xlsx")]
        )
        if not p: return
        try:
            write_excel(rows, p); self._log(f"🧾 Excel yazıldı: {p}")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel yazılamadı: {e}")

    # ---------- Interaktif tablo: sıralama / filtre / kopyala ----------
    def _apply_filter(self):
        txt = nlow(self.tk_filter.get())
        only_incomplete = self.tk_only_incomplete.get()

        # İlk kullanımda tüm iidlere kaydet
        if not self._all_iids:
            self._all_iids = list(self.tree.get_children(""))

        # Hepsini geçici ayır
        for iid in self._all_iids:
            try: self.tree.detach(iid)
            except Exception: pass

        def row_match(vals: List[str]) -> bool:
            # Metin filtresi
            if txt:
                hay = " | ".join([nlow(v) for v in vals])
                if txt not in hay: 
                    return False
            if only_incomplete:
                # Alış tarafında kritik kolonlar: 2..7; Satış tarafında 10..16
                need_cols = [2,3,4,5,6,7, 10,11,12,13,14,15,16]
                if all(nlow(vals[i] if i < len(vals) else "") for i in need_cols):
                    return False
            return True

        for iid in self._all_iids:
            vals = ensure_len(list(self.tree.item(iid, "values")))
            if row_match(vals):
                self.tree.reattach(iid, "", "end")

    def _sort_by(self, col: str):
        # Toggle asc/desc
        asc = not self._sort_state.get(col, True)
        self._sort_state[col] = asc

        def key_func(iid):
            v = self.tree.set(iid, col)
            # Tarih sütunları için YYYY-MM-DD varsayımı
            if col in ("Belge Tarihi","SATIŞ TARİHi"):
                return v or ""
            # Sayı gibi görünenler
            try:
                return float(str(v).replace(".","").replace(",","."))  # TR format toleransı
            except: 
                return v or ""

        iids = list(self.tree.get_children(""))
        iids.sort(key=key_func, reverse=not asc)
        # yeniden sırala
        for idx, iid in enumerate(iids):
            self.tree.move(iid, "", idx)

    def _copy_selection(self, event=None):
        sel = self.tree.selection()
        if not sel: return
        # Seçili satırları tab-separated kopyala (Excel-friendly)
        lines = []
        # Başlıklar:
        lines.append("\t".join(HEADERS))
        for iid in sel:
            vals = [str(x) for x in ensure_len(list(self.tree.item(iid, "values")))]
            lines.append("\t".join(vals))
        txt = "\n".join(lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(txt)
            self._log(f"📋 {len(sel)} satır panoya kopyalandı.")
        except Exception as e:
            self._log(f"📋 Kopyalama hatası: {e}")

    def _popup_menu(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            if iid not in self.tree.selection():
                self.tree.selection_set(iid)
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="Seçili → PDF indir", command=lambda: self._download_selected(mode="pdf"))
            menu.add_command(label="Seçili → XML indir", command=lambda: self._download_selected(mode="xml"))
            menu.add_command(label="Seçili → PDF+XML indir", command=lambda: self._download_selected(mode="both"))
            menu.tk_popup(event.x_root, event.y_root)

# ====================== Çalıştır ======================
if __name__ == "__main__":
    App().mainloop()
