# -*- coding: utf-8 -*-
"""
IMEI → Alış + Satış Birleşik Rapor (NES + GP) – v10.9 (Otomatik Birleşik Arama)

YENİ (v10.9):
- Tek Tuşla Raporlama: "Ara (İndirmeden)" butonu artık çok daha akıllı.
  1. Önce tabloda verilen IMEI listesi için NES API'den tüm alış/satış faturalarını arar ve doldurur.
  2. Ardından, hala alış bilgisi eksik olan IMEI'ler için OTOMATİK olarak
     Gider Pusulası linkini tarar ve kalan boşlukları oradan tamamlar.
- Ana iş akışı basitleştirildi: "IMEI Listesi Yükle" -> "Ara (İndirmeden)". Hepsi bu kadar.
- "Yeni IMEI ekle" kutusu ana iş akışından ayrıldı; "Ara" butonu artık asla yeni IMEI eklemez.
"""

import os, re, io, json, math, threading, csv
from datetime import date
from typing import Any, Dict, List, Optional, Tuple, Set
import xml.etree.ElementTree as ET

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# ====================== Sabitler ======================
DEFAULTS = {
    "api_token": "3B51B4C7C94FF977E42389915CFDA353F6DCE2BF6A2A82C033FBB0950B17CDE8",
    "download_dir": os.getcwd(),
    "out_name": r"C:\Users\siyah\OneDrive\Masaüstü\imei_rapor.xlsx",
    "whitelist_patterns": [
        r"\bYURTIC[IİÇ]|ARAS|MNG|S[ÜU]RAT|PTT|UPS|FEDEX|DHL\b",
        r"\bT[ÜU]RK ?TELEKOM|TTNET|TURKCELL|VODAFONE|LIFECELL|T[ÜU]RKSAT|SUPERONLINE\b",
        r"\bELEKTR[Iİ]K|GAZ|DO[GĞ]AL ?GAZ|ENERJ[İI]|PERAKENDE SAT[Iİ]Ş\b",
        r"\bKULE Y[ÖO]NET[İI]M\b|\bLORAS GAYR[İI]MENKUL\b",
    ],
    "timeout_connect": 15,
    "timeout_read": 90,
    "retries": 4,
    "backoff": 0.6,
}
DEFAULT_DATE_START = "2015-01-01"
def _today_str(): return date.today().strftime("%Y-%m-%d")

# ====================== NES API ======================
EINV_IN_LIST   = "https://api.nes.com.tr/einvoice/v1/incoming/invoices"
EINV_OUT_LIST  = "https://api.nes.com.tr/einvoice/v1/outgoing/invoices"
EARCH_OUT_LIST = "https://api.nes.com.tr/earchive/v1/invoices"

EINV_IN_DOC    = "https://api.nes.com.tr/einvoice/v1/incoming/invoices/{id}"
EINV_OUT_DOC   = "https://api.nes.com.tr/einvoice/v1/outgoing/invoices/{id}"
EARCH_OUT_DOC  = "https://api.nes.com.tr/earchive/v1/invoices/{id}"

PAGE_SIZE = 50
SETTINGS_FILE = "imei_beyanname_v10.json"

NS = {
    "inv": "urn:oasis:names:specification:ubl:schema:xsd:Invoice-2",
    "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
    "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
}

# ---------- TR karakter dönüşümü ----------
TR_MAP = str.maketrans("çğıöşüÇĞİÖŞÜ","cgiosuCGIOSU")
def norm(s: Any) -> str: return ("" if s is None else str(s)).strip()
def nlow(s: Any) -> str: return norm(s).lower().translate(TR_MAP)
def nup(s: Any)  -> str: return norm(s).upper().translate(TR_MAP)

# ---------- Güvenli dosya adı ----------
def safe_filename(name: str) -> str:
    s = norm(name)
    s = re.sub(r'[\\/*?:"<>|]', "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    return (s or "DOSYA")[:120]

# ---------- IMEI yardımcıları ----------
IMEI_RE_STRICT = re.compile(r'(?<!\d)\d{15}(?!\d)')

def _luhn_ok_imei(s15: str) -> bool:
    if len(s15) != 15 or not s15.isdigit(): return False
    total = 0
    for i, ch in enumerate(s15):
        n = int(ch)
        if i % 2 == 1:
            n *= 2
            if n > 9: n -= 9
        total += n
    return (total % 10) == 0

def extract_imeis(text: str) -> List[str]:
    t = norm(text)
    if not t: return []
    out = set()
    for m in IMEI_RE_STRICT.finditer(t):
        s = m.group(0)
        if _luhn_ok_imei(s):
            out.add(s)
    return sorted(out)

# ---------- Marka normalizasyonu ----------
def _cp(p): return re.compile(p, re.I)
_BRAND_PATTERNS: List[Tuple[str, re.Pattern]] = [
    ("APPLE",   _cp(r"\bAPPLE\b|\bIPHONE?\b|\bAPLE\b|\bI ?PHONE\b")),
    ("SAMSUNG", _cp(r"\bSAMSUNG\b|\bGALAXY\b|\bSM[-\s]")),
    ("XIAOMI",  _cp(r"\bXIAOM[Iİ]\b|\bREDMI\b|\bPOCO\b|\bMI[-\s]")),
    ("HUAWEI",  _cp(r"\bHUAWEI\b|\bHUAWE\b|\bP\d{2}\b|\bMATE\b(?!.*HONOR)")),
    ("HONOR",   _cp(r"\bHONOR\b")),
    ("OPPO",    _cp(r"\bOPPO\b")),
    ("REALME",  _cp(r"\bREALME\b")),
    ("VIVO",    _cp(r"\bVIVO\b")),
    ("TECNO",   _cp(r"\bTECNO\b")),
    ("NOKIA",   _cp(r"\bNOKIA\b")),
    ("CASPER",  _cp(r"\bCASPER\b")),
    ("GENERAL MOBILE", _cp(r"\bGENERAL\s*MOBILE\b|\bGM\s?\d+\b")),
    ("INFINIX", _cp(r"\bINFINIX\b")),
    ("REEDER",  _cp(r"\bREEDER\b")),
]
def brand_from_text(t: str) -> str:
    U = nup(t)
    for brand, pat in _BRAND_PATTERNS:
        if pat.search(U):
            if brand == "HUAWEI" and "HONOR" in U:
                continue
            return brand
    return "Bilinmeyen"

# ---------- Metin ipuçları ----------
KEY_2EL  = re.compile(r"\b(2\.?\s*EL|İKİNCİ\s*EL|IKINCI\s*EL|SECOND\s*HAND)\b", re.I)
KEY_REF  = re.compile(r"\b(YENİLENMİŞ|YENILENMIS|REFURB)\b", re.I)

# ---------- Belge no desenleri ----------
DOCNO_RE = re.compile(r"\bE(?:AR|FR)\d{13}\b", re.I)

# ---------- Ağ / HTTP ----------
TIMEOUT_CONNECT = DEFAULTS["timeout_connect"]
TIMEOUT_READ    = DEFAULTS["timeout_read"]
RETRIES         = DEFAULTS["retries"]
BACKOFF         = DEFAULTS["backoff"]
SESSION = None

def make_session(retries: int = None, backoff: float = None) -> requests.Session:
    r = retries if retries is not None else RETRIES
    b = backoff if backoff is not None else BACKOFF
    s = requests.Session()
    retry = Retry(
        total=r, connect=r, read=r, status=r,
        backoff_factor=b,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET", "POST"]),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=100, pool_maxsize=100)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s

def get_session() -> requests.Session:
    global SESSION
    if SESSION is None:
        SESSION = make_session()
    return SESSION

def headers(token: str, typ="json") -> Dict[str, str]:
    h = {
        "Authorization": f"Bearer {token.strip()}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": "IMEI-NES-Client/10.9",
    }
    if typ == "xml": h["Accept"] = "application/xml"
    if typ == "pdf": h["Accept"]  = "application/pdf"
    if typ == "bin": h["Accept"]  = "*/*"
    return h

def http_get(url: str, *, token: Optional[str], typ="json", params=None, log=None, stop_evt: Optional[threading.Event]=None):
    if stop_evt is not None and stop_evt.is_set():
        return None
    try:
        sess = get_session()
        hdrs = headers(token, typ) if token is not None else {"User-Agent": "IMEI-NES-Client/10.9"}
        r = sess.get(url, headers=hdrs, params=params, timeout=(TIMEOUT_CONNECT, TIMEOUT_READ))
        if r.status_code == 200:
            return r
        if log: log(f"[HTTP] Hata {r.status_code}: {url} → {r.text[:300]}")
        return None
    except requests.Timeout:
        if log: log(f"[HTTP] Zaman aşımı: {url}")
        return None
    except requests.RequestException as e:
        if log: log(f"[HTTP] İstek hatası: {e}")
        return None

def xfind(t: ET.Element, path: str) -> Optional[ET.Element]: return t.find(path, NS) if t is not None else None
def xtext(t: ET.Element, path: str) -> str:
    el = xfind(t, path); return norm(el.text) if el is not None else ""

# ---------- KDV yardımcıları ----------
def _to_float(x: str) -> Optional[float]:
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return None

def _mode_or_first(vals: List[float]) -> Optional[float]:
    if not vals: return None
    try:
        from statistics import mode
        return float(mode(vals))
    except Exception:
        return float(vals[0])

# ====================== XML Parser ======================
class Parsed:
    def __init__(self):
        self.invoice_no = ""; self.issue_date = ""
        self.payable = ""; self.tax_total = ""
        self.buyer_name = ""; self.buyer_id = ""; self.buyer_id_type = ""
        self.supplier_name = ""; self.supplier_id = ""; self.supplier_id_type = ""
        self.description = ""; self.items: List[str] = []
        self.imeis: List[str] = []; self.brand = ""; self.model = ""
        self.lines: List[Dict[str,Any]] = []
        self.inv_kdv: Optional[float] = None
        self.text_upper: str = ""

def _line_kdv_percent(line: ET.Element) -> Optional[float]:
    p1 = xfind(line, "cac:TaxTotal/cac:TaxSubtotal/cbc:Percent")
    if p1 is not None and norm(p1.text):
        f = _to_float(p1.text)
        if f is not None: return f
    p2 = xfind(line, "cac:Item/cac:ClassifiedTaxCategory/cbc:Percent")
    if p2 is not None and norm(p2.text):
        f = _to_float(p2.text)
        if f is not None: return f
    p3 = xfind(line, ".//cbc:Percent")
    if p3 is not None and norm(p3.text):
        f = _to_float(p3.text);
        if f is not None: return f
    return None

def _invoice_kdv_candidates(root: ET.Element) -> List[float]:
    cands: List[float] = []
    for el in root.findall(".//cac:TaxTotal/cac:TaxSubtotal/cbc:Percent", NS):
        f = _to_float(el.text)
        if f is not None: cands.append(f)
    if not cands:
        for el in root.findall(".//cbc:Percent", NS):
            f = _to_float(el.text)
            if f in (1.0, 8.0, 10.0, 18.0, 20.0): cands.append(f)
    return cands

def parse_invoice_xml(xml_bytes: bytes) -> Parsed:
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return Parsed()
    P = Parsed()
    P.invoice_no  = xtext(root, "cbc:ID")
    P.issue_date  = xtext(root, "cbc:IssueDate")
    P.tax_total   = xtext(root, "cac:TaxTotal/cbc:TaxAmount")
    P.payable     = xtext(root, "cac:LegalMonetaryTotal/cbc:PayableAmount")
    cust = xfind(root, "cac:AccountingCustomerParty/cac:Party")
    if xfind(cust, "cac:Person") is not None:
        P.buyer_name = (xtext(cust, "cac:Person/cbc:FirstName")+" "+xtext(cust, "cac:Person/cbc:FamilyName")).strip()
    else:
        P.buyer_name = xtext(cust, "cac:PartyName/cbc:Name")
    bid = xfind(cust, "cac:PartyIdentification/cbc:ID")
    P.buyer_id = norm(bid.text) if bid is not None else ""
    P.buyer_id_type = bid.attrib.get("schemeID","") if bid is not None else ""
    if not P.buyer_id_type:
        if re.fullmatch(r"\d{11}", P.buyer_id): P.buyer_id_type = "TCKN"
        elif re.fullmatch(r"\d{10}", P.buyer_id): P.buyer_id_type = "VKN"
    sup = xfind(root, "cac:AccountingSupplierParty/cac:Party")
    P.supplier_name = xtext(sup, "cac:PartyName/cbc:Name")
    sid = xfind(sup, "cac:PartyIdentification/cbc:ID")
    P.supplier_id = norm(sid.text) if sid is not None else ""
    P.supplier_id_type = sid.attrib.get("schemeID","") if sid is not None else ""
    if not P.supplier_id_type:
        if re.fullmatch(r"\d{11}", P.supplier_id): P.supplier_id_type = "TCKN"
        elif re.fullmatch(r"\d{10}", P.supplier_id): P.supplier_id_type = "VKN"
    notes = [norm(n.text) for n in root.findall("cbc:Note", NS) if norm(n.text)]
    P.description = " | ".join(notes)
    kdv_seen: List[float] = []
    for line in root.findall("cac:InvoiceLine", NS):
        name = xtext(line, "cac:Item/cbc:Name")
        desc = xtext(line, "cac:Item/cbc:Description")
        props = [norm(p.text) for p in line.findall("cac:Item/cac:AdditionalItemProperty/cbc:Value", NS)]
        blob = " ".join([name, desc] + [p for p in props if p]).strip()
        if not blob: continue
        kdv = _line_kdv_percent(line)
        if kdv is not None: kdv_seen.append(kdv)
        P.lines.append({
            "blob": blob, "unit_price": xtext(line, "cac:Price/cbc:PriceAmount"),
            "line_total": xtext(line, "cac:LineExtensionAmount"),
            "qty": xtext(line, "cbc:InvoicedQuantity"), "kdv": kdv
        })
        P.items.append(blob)
    P.inv_kdv = _mode_or_first(kdv_seen or _invoice_kdv_candidates(root))
    hay = " \n ".join([P.description] + P.items)
    P.imeis = extract_imeis(hay)
    P.brand = brand_from_text(hay)
    P.model = P.lines[0]["blob"] if P.lines else (P.items[0] if P.items else "")
    P.text_upper = nup(hay)
    return P

# ====================== Listeleme / Yardımcılar ======================
def paged_list(url: str, token: str, start: str, end: str, log, stop_evt, archived: Optional[bool]=None, section_name: str="") -> List[Dict[str,Any]]:
    if not start and not end:
        start, end = DEFAULT_DATE_START, _today_str()
        log(f"[{section_name}] Tarih boş → {start}..{end} aralığı kullanılacak.")
    out, page, total = [], 1, None
    while True:
        if stop_evt.is_set(): break
        params = {"sort":"createdAt desc", "page":page, "pageSize":PAGE_SIZE}
        if archived is not None: params["archived"] = "true" if archived else "false"
        if start: params["startDate"] = f"{start}T00:00:00+03:00"
        if end:   params["endDate"]   = f"{end}T23:59:59+03:00"
        r = http_get(url, token=token, typ="json", params=params, log=log, stop_evt=stop_evt)
        if r is None:
            log(f"[{section_name}] İstek başarısız/timeout. Döngü sonlandırıldı.")
            break
        try: data = r.json() or {}
        except ValueError:
            log(f"[{section_name}] JSON çözümlenemedi."); break
        if total is None:
            tc = data.get("totalCount") or 0
            total = max(1, math.ceil(tc / PAGE_SIZE)) if isinstance(tc, int) else 1
        batch = data.get("data") or data.get("invoices") or []
        if not batch: break
        log(f"[{section_name}] Sayfa {page}/{total} → {len(batch)} kayıt")
        out.extend(batch)
        if page >= total: break
        page += 1
    return out

def list_both_archived(url: str, token: str, start: str, end: str, log, stop_evt, section_name: str) -> List[Dict[str,Any]]:
    items = []
    log(f"[{section_name}] Arşivsiz çekiliyor...")
    items += paged_list(url, token, start, end, log, stop_evt, archived=False, section_name=section_name)
    log(f"[{section_name}] Arşivli çekiliyor...")
    items += paged_list(url, token, start, end, log, stop_evt, archived=True,  section_name=section_name)
    seen, uniq = set(), []
    for m in items:
        mid = str(m.get("id") or "")
        if mid and mid not in seen:
            seen.add(mid); uniq.append(m)
    log(f"[{section_name}] Birleştirildi (tekil): {len(uniq)} kayıt")
    return uniq

def fetch_xml_by(url_tpl: str, token: str, inv_id: str) -> Optional[bytes]:
    r = http_get(f"{url_tpl.format(id=inv_id)}/xml", token=token, typ="xml")
    return (r.content if r is not None else None)

def fetch_pdf_by(url_tpl: str, token: str, inv_id: str) -> Optional[bytes]:
    r = http_get(f"{url_tpl.format(id=inv_id)}/pdf", token=token, typ="pdf")
    return (r.content if r is not None else None)

# ====================== Ayarlar / Excel ======================
HEADERS = [
    "imei","Bulunma","Tck/Vkn","Belge Türü","Belge Tarihi","Belge No","Alınan Kişi","Borç Tutar",
    "Marka","MODEL",
    "SATIŞ TARİHi","ALICI ADI SOYADI","SATIŞ BEDELİ","KDV TUTARI","SATIŞ BELGESİNİN NUMARASI",
    "ALICI KİMLİK TÜRÜ","ALICI KİMLİK NO",
    "Sütun1","ALIŞ BELGELERİ TÜRÜ","DURUMU",
    "ALIŞ KDV","SATIŞ KDV","SINIF","GEREKÇE"
]
def ensure_len(vals: List[Any]) -> List[Any]:
    if len(vals) < len(HEADERS):
        vals = vals + [""]*(len(HEADERS)-len(vals))
    elif len(vals) > len(HEADERS):
        vals = vals[:len(HEADERS)]
    return vals

def load_settings() -> Dict[str,Any]:
    s = DEFAULTS.copy()
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f: s.update(json.load(f))
        except Exception: pass
    try:
        s["_whitelist_compiled"] = [re.compile(pat, re.I) for pat in s.get("whitelist_patterns",[])]
    except Exception:
        s["_whitelist_compiled"] = []
    return s
def save_settings(s: Dict[str,Any]):
    s2 = {k:v for k,v in s.items() if k != "_whitelist_compiled"}
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f: json.dump(s2, f, ensure_ascii=False, indent=2)

def write_excel(rows: List[List[Any]], out_path: str):
    wb = Workbook(); ws = wb.active; ws.title = "IMEI_RAPOR"
    ws.append(HEADERS)
    for r in rows: ws.append(ensure_len(r))
    for col in range(1, len(HEADERS)+1):
        mx = max((len(str(ws.cell(row=i, column=col).value or "")) for i in range(1, ws.max_row+1)), default=12)
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, mx+2), 60)
    wb.save(out_path)

def is_whitelisted_supplier(name: str, settings: Dict[str,Any]) -> bool:
    if not name: return False
    U = nup(name)
    for pat in settings.get("_whitelist_compiled", []):
        if pat.search(U): return True
    return False

# ====================== GP (Gider Pusulası) – Okuyucu ======================
def _norm_header(h: Any) -> str: return nlow(h).replace("  ", " ").strip()
_HEADER_ALIAS = {
    "imei": "imei", "imei/seri": "imei", "seri no": "imei", "serino": "imei",
    "tck/vkn": "Tck/Vkn", "tckn": "Tck/Vkn", "vkn": "Tck/Vkn", "tc kimlik": "Tck/Vkn",
    "belge turu": "Belge Türü", "belge tarihi": "Belge Tarihi", "belge no": "Belge No",
    "alinan kisi": "Alınan Kişi", "alınan kişi": "Alınan Kişi", "borc tutar": "Borç Tutar",
    "marka": "Marka", "model": "MODEL", "açıklama": "MODEL", "aciklama": "MODEL", "urun": "MODEL", "ürün": "MODEL",
    "satis tarihi": "SATIŞ TARİHi", "alici adi soyadi": "ALICI ADI SOYADI",
    "satis bedeli": "SATIŞ BEDELİ", "kdv tutari": "KDV TUTARI",
    "satis belgesinin numarasi": "SATIŞ BELGESİNİN NUMARASI",
    "alici kimlik turu": "ALICI KİMLİK TÜRÜ", "alici kimlik no": "ALICI KİMLİK NO",
    "sutun1": "Sütun1", "şube": "Sütun1", "sube": "Sütun1", "magaza":"Sütun1",
    "alis belgeleri turu": "ALIŞ BELGELERİ TÜRÜ", "durumu": "DURUMU",
}
def _build_header_map(ws) -> Optional[Dict[str, Any]]:
    for r in range(1, min(11, ws.max_row)+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        normed = [_norm_header(x) for x in row]
        col_to_name = {}; hit = 0; has_imei = False
        for idx, hh in enumerate(normed):
            if not hh: continue
            key = hh.replace("ı","i"); key = re.sub(r"[^a-z0-9/ ]", "", key)
            name = _HEADER_ALIAS.get(key, None)
            if name:
                col_to_name[idx] = name; hit += 1
                if name == "imei": has_imei = True
        if has_imei and hit >= 5: return {"row": r, "cols": col_to_name}
    return None
def parse_gp_template_workbook(wb, log) -> List[List[Any]]:
    out_rows: List[List[Any]] = []
    for ws in wb.worksheets:
        hm = _build_header_map(ws)
        if not hm: continue
        hdr_row = hm["row"]; colmap = hm["cols"]; taken = 0
        for r in range(hdr_row+1, ws.max_row+1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
            if not any(norm(v) for v in vals): continue
            row_dict = {name: "" for name in HEADERS}
            for cidx, name in colmap.items():
                if cidx < len(vals): row_dict[name] = norm(vals[cidx])
            imei_val = row_dict.get("imei", "")
            if not re.fullmatch(r"\d{15}", imei_val or ""):
                merged = " | ".join([norm(v) for v in vals])
                m_all = extract_imeis(merged)
                if not m_all: continue
                row_dict["imei"] = m_all[0]
            if nup(row_dict.get("Belge Türü","")) in ("GMA","GIDER PUSULASI","GİDER PUSULASI"):
                row_dict["Belge Türü"] = "GİDER PUSULASI"
                row_dict["ALIŞ BELGELERİ TÜRÜ"] = row_dict.get("ALIŞ BELGELERİ TÜRÜ") or "Gider Pusulası"
            if not row_dict.get("Marka"):
                row_dict["Marka"] = brand_from_text(row_dict.get("MODEL",""))
            if any(row_dict.get(k) for k in ("SATIŞ TARİHi","ALICI ADI SOYADI","SATIŞ BEDELİ","SATIŞ BELGESİNİN NUMARASI")):
                row_dict["DURUMU"] = row_dict.get("DURUMU") or "Satılmış"
            else:
                row_dict["DURUMU"] = row_dict.get("DURUMU") or "Satılabilir"
            row_dict["Bulunma"] = row_dict.get("Bulunma") or "GP"
            out_rows.append([row_dict[h] for h in HEADERS]); taken += 1
        log(f"[GP] Sayfa '{ws.title}': {taken} satır alındı.")
    return out_rows
GP_KW = { "imei": ["imei","seri","serial","serı","serino","seri no","seri-no"], "tarih": ["tarih","alış tarihi","alis tarihi","işlem tarihi","islem tarihi","gp tarihi"], "bedel": ["tutar","bedel","fiyat","ödenen","odenen","ucret","ücret"], "ad": ["adı","ad soyad","adsoyad","satıcı","satici","müsteri","musteri","alan","veren"], "sube": ["şube","sube","magaza","mağaza"], "aciklama": ["açıklama","aciklama","not","ürün","urun","cihaz","model"] }
def _find_col_idx(headers_row: List[Any], keys: List[str]) -> Optional[int]:
    for i, h in enumerate(headers_row):
        hh = nlow(h)
        for k in keys:
            if k in hh: return i
    return None
def _scan_header(ws) -> Tuple[int, Dict[str,int]]:
    for r in range(1, min(12, ws.max_row)+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        idx_imei = _find_col_idx(row, GP_KW["imei"])
        if idx_imei is not None:
            cols = { "imei": idx_imei, "tarih": _find_col_idx(row, GP_KW["tarih"]) or -1, "bedel": _find_col_idx(row, GP_KW["bedel"]) or -1, "ad": _find_col_idx(row, GP_KW["ad"]) or -1, "sube": _find_col_idx(row, GP_KW["sube"]) or -1, "aciklama": _find_col_idx(row, GP_KW["aciklama"]) or -1, }
            return r, cols
    return 1, {"imei": 0, "tarih": -1, "bedel": -1, "ad": -1, "sube": -1, "aciklama": -1}
def parse_gp_workbook(wb, log) -> List[Dict[str,Any]]:
    out = []
    for ws in wb.worksheets:
        head_row, cols = _scan_header(ws)
        start = head_row + 1; found_rows = 0
        for r in range(start, ws.max_row+1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
            text_row = " | ".join([norm(v) for v in vals])
            imeis = extract_imeis(text_row)
            if not imeis and cols["imei"] >= 0:
                v = norm(vals[cols["imei"]] if cols["imei"] < len(vals) else "")
                if re.fullmatch(r"\d{15}", v) and _luhn_ok_imei(v): imeis = [v]
            if not imeis: continue
            for im in imeis:
                item = { "imei": im, "tarih": norm(vals[cols["tarih"]]) if cols["tarih"] >= 0 else "", "bedel": norm(vals[cols["bedel"]]) if cols["bedel"] >= 0 else "", "ad": norm(vals[cols["ad"]]) if cols["ad"] >= 0 else "", "sube": norm(vals[cols["sube"]]) if cols["sube"] >= 0 else "", "aciklama": norm(vals[cols["aciklama"]]) if cols["aciklama"] >= 0 else "", }
                out.append(item); found_rows += 1
        log(f"[GP] Sayfa '{ws.title}': {found_rows} satır/IMEI çıkarıldı.")
    return out

# ====================== GUI ======================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("IMEI → Alış + Satış Birleşik Rapor (NES + GP) – v10.9")
        self.geometry("1540x1000")
        self.settings = load_settings()
        self.stop_evt = threading.Event()
        self.worker = None
        self.rows: List[List[Any]] = []
        self.iid_to_row_index: Dict[str,int] = {}
        self.imei_to_iid: Dict[str,str] = {}
        self.iid_to_ids: Dict[str, Dict[str,str]] = {}
        self.seen_in_pairs  = set(); self.seen_out_pairs = set()
        self.imei_first_doc_in: Dict[str,str] = {}
        self.imei_first_doc_out: Dict[str,str] = {}
        self.force_imeis_order: List[str] = []
        self.force_imeis_set: set = set()
        self.imei_kdv_in: Dict[str, Set[int]]  = {}
        self.imei_kdv_out: Dict[str, Set[int]] = {}
        self.imei_flags: Dict[str, Dict[str,bool]] = {}
        self.docnos_filter: Set[str] = set()
        self._build_ui()

    def _build_ui(self):
        top = ttk.LabelFrame(self, text="Kimlik & Klasör"); top.pack(fill="x", padx=12, pady=6)
        ttk.Label(top, text="API Token:").grid(row=0, column=0, sticky="e")
        self.tk_token = tk.StringVar(value=self.settings.get("api_token",""))
        ttk.Entry(top, textvariable=self.tk_token, show="•", width=60).grid(row=0, column=1, columnspan=3, sticky="we", padx=6)
        ttk.Button(top, text="Kaydet", command=self._save_settings).grid(row=0, column=4, padx=6)
        ttk.Label(top, text="İndirme Klasörü:").grid(row=0, column=5, sticky="e")
        self.tk_dir = tk.StringVar(value=self.settings.get("download_dir", os.getcwd()))
        ttk.Entry(top, textvariable=self.tk_dir, width=40).grid(row=0, column=6, sticky="we")
        ttk.Button(top, text="Seç", command=self._pick_dir).grid(row=0, column=7, padx=6)
        for c in range(8): top.columnconfigure(c, weight=1)
        flt = ttk.LabelFrame(self, text="Filtreler (Sadece NES API için geçerlidir)"); flt.pack(fill="x", padx=12, pady=6)
        for c in range(16): flt.columnconfigure(c, weight=1)
        ttk.Label(flt, text="Tedarikçi Ünvan içerir:").grid(row=0, column=0, sticky="e")
        self.tk_unvan = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_unvan, width=26).grid(row=0, column=1, sticky="w")
        ttk.Label(flt, text="TCKN:").grid(row=0, column=2, sticky="e")
        self.tk_tckn = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_tckn, width=14).grid(row=0, column=3, sticky="w")
        ttk.Label(flt, text="VKN:").grid(row=0, column=4, sticky="e")
        self.tk_vkn = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_vkn, width=14).grid(row=0, column=5, sticky="w")
        ttk.Separator(flt, orient="vertical").grid(row=0, column=6, rowspan=2, sticky="ns", padx=8)
        ttk.Label(flt, text="Alıcı Ünvan içerir:").grid(row=0, column=7, sticky="e")
        self.tk_alici_unvan = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_alici_unvan, width=26).grid(row=0, column=8, sticky="w")
        ttk.Label(flt, text="Alıcı TCKN:").grid(row=0, column=9, sticky="e")
        self.tk_alici_tckn = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_alici_tckn, width=14).grid(row=0, column=10, sticky="w")
        ttk.Label(flt, text="Alıcı VKN:").grid(row=0, column=11, sticky="e")
        self.tk_alici_vkn = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_alici_vkn, width=14).grid(row=0, column=12, sticky="w")
        self.tk_use_date = tk.BooleanVar(value=False)
        ttk.Checkbutton(flt, text="Tarih filtresi", variable=self.tk_use_date).grid(row=0, column=13, sticky="w")
        ttk.Label(flt, text="Başlangıç (YYYY-MM-DD):").grid(row=0, column=14, sticky="e")
        self.tk_start = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_start, width=12).grid(row=0, column=15, sticky="w")
        ttk.Label(flt, text="Bitiş:").grid(row=1, column=14, sticky="e")
        self.tk_end = tk.StringVar(); ttk.Entry(flt, textvariable=self.tk_end, width=12).grid(row=1, column=15, sticky="w")
        opts = ttk.LabelFrame(self, text="Arama Opsiyonları"); opts.pack(fill="x", padx=12, pady=6)
        self.tk_terms = tk.Text(opts, height=3); self.tk_terms.pack(fill="x", padx=8, pady=6)
        sub = ttk.Frame(opts); sub.pack(fill="x", padx=8, pady=4)
        self.tk_scan_sales = tk.BooleanVar(value=True); self.tk_only_match = tk.BooleanVar(value=True); self.tk_get_pdf = tk.BooleanVar(value=False); self.tk_get_xml = tk.BooleanVar(value=True); self.tk_auto_xlsx = tk.BooleanVar(value=True)
        self.tk_add_new_imeis = tk.BooleanVar(value=False)
        ttk.Checkbutton(sub, text="Bulunan yeni IMEI'leri listeye ekle", variable=self.tk_add_new_imeis).pack(side="left", padx=6)
        ttk.Separator(sub, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Checkbutton(sub, text="Satışları tara (e-Arşiv + Giden)", variable=self.tk_scan_sales).pack(side="left", padx=6)
        ttk.Checkbutton(sub, text="PDF indir", variable=self.tk_get_pdf).pack(side="left", padx=6)
        ttk.Checkbutton(sub, text="XML indir", variable=self.tk_get_xml).pack(side="left", padx=6)
        ttk.Checkbutton(sub, text="Bittiğinde otomatik Excel yaz", variable=self.tk_auto_xlsx).pack(side="left", padx=6)
        docf = ttk.LabelFrame(self, text="Fatura No Listesi (EAR/EFR) – isteğe bağlı"); docf.pack(fill="x", padx=12, pady=6)
        row = ttk.Frame(docf); row.pack(fill="x", padx=8, pady=4)
        self.tk_docnos_count = tk.StringVar(value="Seçili fatura: 0")
        ttk.Button(row, text="TXT'den Yükle", command=self._load_docnos_from_file).pack(side="left", padx=4)
        ttk.Button(row, text="Panodan Yapıştır", command=self._paste_docnos_clip).pack(side="left", padx=4)
        ttk.Button(row, text="Temizle", command=self._clear_docnos).pack(side="left", padx=4)
        ttk.Label(row, textvariable=self.tk_docnos_count).pack(side="left", padx=10)
        net = ttk.LabelFrame(opts, text="Ağ (timeout & retry)"); net.pack(fill="x", padx=8, pady=6)
        self.tk_timeout_c = tk.IntVar(value=int(self.settings.get("timeout_connect", DEFAULTS["timeout_connect"]))); self.tk_timeout_r = tk.IntVar(value=int(self.settings.get("timeout_read", DEFAULTS["timeout_read"]))); self.tk_retries   = tk.IntVar(value=int(self.settings.get("retries", DEFAULTS["retries"]))); self.tk_backoff   = tk.DoubleVar(value=float(self.settings.get("backoff", DEFAULTS["backoff"])))
        ttk.Label(net, text="Bağlantı timeout (sn):").grid(row=0, column=0, sticky="e"); ttk.Entry(net, textvariable=self.tk_timeout_c, width=6).grid(row=0, column=1, sticky="w", padx=6); ttk.Label(net, text="Okuma timeout (sn):").grid(row=0, column=2, sticky="e"); ttk.Entry(net, textvariable=self.tk_timeout_r, width=6).grid(row=0, column=3, sticky="w", padx=6); ttk.Label(net, text="Retry deneme:").grid(row=0, column=4, sticky="e"); ttk.Entry(net, textvariable=self.tk_retries, width=4).grid(row=0, column=5, sticky="w", padx=6); ttk.Label(net, text="Backoff (saniye çarpanı):").grid(row=0, column=6, sticky="e"); ttk.Entry(net, textvariable=self.tk_backoff, width=6).grid(row=0, column=7, sticky="w", padx=6)
        for c in range(8): net.columnconfigure(c, weight=1)
        ext = ttk.LabelFrame(self, text="Gider Pusulası Kaynakları (URL veya Dosya)"); ext.pack(fill="x", padx=12, pady=6)
        self.tk_gp_urls = tk.Text(ext, height=3); self.tk_gp_urls.pack(fill="x", padx=6, pady=6)
        self.tk_gp_urls.insert("1.0", "https://docs.google.com/spreadsheets/d/e/2PACX-1vSDMPeXeKs0HSD38CJst-_1AevO_YuZYtQa7jg-ra0OWQmxi-6qqXGEbqDO_I8ToQ/pub?output=xlsx\n")
        btns = ttk.Frame(ext); btns.pack(fill="x")
        ttk.Button(btns, text="Manuel URL'den Yükle/Güncelle (GP)", command=self._load_gp_from_urls).pack(side="left", padx=6, pady=4)
        ttk.Button(btns, text="Manuel Dosyadan Yükle/Güncelle (GP)", command=self._load_gp_from_file).pack(side="left", padx=6, pady=4)
        act = ttk.Frame(self); act.pack(fill="x", padx=12, pady=4)
        ttk.Button(act, text="1. IMEI Listesi Yükle", command=self._load_imei_list).pack(side="left", padx=4)
        ttk.Button(act, text="2. Raporu Tamamla (NES + GP)", command=self._start_scan).pack(side="left", padx=4)
        ttk.Button(act, text="Seçili → İndir", command=self._download_selected).pack(side="left", padx=4)
        ttk.Button(act, text="Excel'e Aktar", command=self._export_excel).pack(side="left", padx=4)
        self.btn_stop = ttk.Button(act, text="Durdur", command=self._stop_now, state="disabled"); self.btn_stop.pack(side="right", padx=4)
        self.tree = ttk.Treeview(self, columns=HEADERS, show="headings", height=22)
        for c in HEADERS:
            self.tree.heading(c, text=c); w = 150
            if c in ("Marka","MODEL","GEREKÇE"): w = 320
            if c in ("Borç Tutar","SATIŞ BEDELİ","KDV TUTARI"): w = 120
            if c in ("Sütun1",): w = 320
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=12, pady=6)
        ysb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview); self.tree.configure(yscrollcommand=ysb.set); ysb.place(in_=self.tree, relx=1.0, rely=0, relheight=1.0, x=-16)
        logf = ttk.LabelFrame(self, text="Log"); logf.pack(fill="x", padx=12, pady=6)
        self.log = scrolledtext.ScrolledText(logf, height=11); self.log.pack(fill="x", padx=8, pady=6)

    def _log(self, msg: str): self.log.insert(tk.END, msg + "\n"); self.log.see(tk.END)
    def _save_settings(self):
        s = self.settings; s["api_token"] = self.tk_token.get().strip(); s["download_dir"] = self.tk_dir.get().strip() or os.getcwd(); s["out_name"] = s.get("out_name", DEFAULTS["out_name"]); s["timeout_connect"] = int(self.tk_timeout_c.get()); s["timeout_read"]    = int(self.tk_timeout_r.get()); s["retries"] = int(self.tk_retries.get()); s["backoff"] = float(self.tk_backoff.get()); save_settings(s); messagebox.showinfo("Bilgi", "Ayarlar kaydedildi.")
    def _pick_dir(self):
        d = filedialog.askdirectory(title="İndirme klasörü")
        if d: self.tk_dir.set(d)
    def _stop_now(self): self.stop_evt.set(); self._log("❌ Durdurma istendi. Mevcut sayfa bitince durur."); self.btn_stop.config(state="disabled")
    def _extract_docnos(self, text: str) -> List[str]:
        if not text: return []
        seen = set(); out = []
        for m in DOCNO_RE.findall(text.upper()):
            if m not in seen: seen.add(m); out.append(m)
        return out
    def _set_docnos(self, arr: List[str]):
        self.docnos_filter = set(arr or []); self.tk_docnos_count.set(f"Seçili fatura: {len(self.docnos_filter)}")
        if arr: self._log(f"[FILTRE] {len(arr)} EAR/EFR fatura yüklendi.")
        else: self._log("[FILTRE] Fatura listesi temizlendi.")
    def _load_docnos_from_file(self):
        p = filedialog.askopenfilename(title="Fatura Listesi (.txt)", filetypes=[("Metin","*.txt"),("Tümü","*.*")])
        if not p: return
        try:
            with open(p,"r",encoding="utf-8") as f: arr = self._extract_docnos(f.read())
            self._set_docnos(arr)
        except Exception as e: messagebox.showerror("Hata", f"Dosya okunamadı: {e}")
    def _paste_docnos_clip(self):
        try: txt = self.clipboard_get()
        except Exception: messagebox.showerror("Hata", "Panoda veri yok."); return
        arr = self._extract_docnos(txt)
        if not arr: messagebox.showinfo("Bilgi", "Panodaki metinde EAR/EFR bulunamadı.")
        self._set_docnos(arr)
    def _clear_docnos(self): self._set_docnos([])
    def _load_imei_list(self):
        p = filedialog.askopenfilename(title="IMEI listesi seç (Excel/CSV/TXT)", filetypes=[("Excel","*.xlsx *.xls"),("CSV","*.csv"),("Metin","*.txt"),("Tümü","*.*")])
        if not p: return
        self.tree.delete(*self.tree.get_children())
        self.rows.clear(); self.iid_to_row_index.clear(); self.imei_to_iid.clear(); self.iid_to_ids.clear()
        self.force_imeis_order.clear(); self.force_imeis_set.clear()

        imeis_from_file: List[Dict] = []
        try:
            ext = os.path.splitext(p)[1].lower()
            if ext in (".xlsx",".xls"):
                wb = load_workbook(p, data_only=True)
                template_rows = parse_gp_template_workbook(wb, self._log)
                if template_rows:
                    for r_list in template_rows:
                        row_dict = {h: (r_list[i] if i < len(r_list) else "") for i, h in enumerate(HEADERS)}
                        if row_dict.get("imei"): imeis_from_file.append(row_dict)
                else: # Basit liste, sadece ilk sütunu oku
                    ws = wb.active
                    for r in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                        v = str(r[0]).strip() if r and r[0] is not None else ""
                        if re.fullmatch(r"\d{15}", v) and _luhn_ok_imei(v): imeis_from_file.append({"imei": v})
            else: # CSV/TXT
                with open(p, "r", encoding="utf-8-sig") as f:
                    content = f.read()
                    for v in extract_imeis(content):
                        imeis_from_file.append({"imei": v})
        except Exception as e: messagebox.showerror("Hata", f"IMEI listesi okunamadı: {e}"); return

        for item_dict in imeis_from_file:
            im = item_dict["imei"]
            if im in self.force_imeis_set: continue
            self.force_imeis_order.append(im); self.force_imeis_set.add(im)
            row = [item_dict.get(h, "") for h in HEADERS]
            iid = self.tree.insert("", "end", values=ensure_len(row))
            self.iid_to_row_index[iid] = len(self.rows); self.rows.append(ensure_len(row)); self.imei_to_iid[im] = iid

        self._log(f"📥 IMEI listesi yüklendi ve tablo oluşturuldu: {len(self.force_imeis_set)} adet")
    def _merge_gp_ready_rows(self, rows_ready: List[List[Any]]):
        add_new = self.tk_add_new_imeis.get()
        added = merged = skipped = 0
        for ready in rows_ready:
            ready = ensure_len(ready)
            rd = {h: (ready[i] if i < len(ready) else "") for i, h in enumerate(HEADERS)}
            im = rd["imei"]
            if not im: continue
            self._mark_flags(im, ref=bool(KEY_REF.search(nup(rd.get("MODEL","")))), is2=bool(KEY_2EL.search(nup(rd.get("MODEL","")))), gp=True)
            iid = self.imei_to_iid.get(im)
            if iid:
                cur = ensure_len(list(self.tree.item(iid, "values")))
                for i, h in enumerate(HEADERS[:20]):
                    if not cur[i] and rd.get(h): cur[i] = rd[h]
                cur[1] = "XML+GP" if (cur[1] and cur[1] != "GP") else (cur[1] or "GP")
                self.tree.item(iid, values=cur)
                idxr = self.iid_to_row_index.get(iid)
                if idxr is not None: self.rows[idxr] = ensure_len(cur)
                merged += 1
            else:
                if add_new:
                    if im not in self.force_imeis_set:
                        self.force_imeis_order.append(im); self.force_imeis_set.add(im)
                    row = [rd.get(h, "") for h in HEADERS]
                    iid = self.tree.insert("", "end", values=ensure_len(row))
                    self.iid_to_row_index[iid] = len(self.rows); self.rows.append(ensure_len(row)); self.imei_to_iid[im] = iid; added += 1
                else: skipped += 1
            if self.imei_to_iid.get(im):
                self._update_classification_for(im)
        self._log(f"[GP Şablon] Birleştirilen={merged}, Yeni Eklenen={added}, Atlanan={skipped}")
    def _load_gp_from_urls(self):
        urls = [u.strip() for u in self.tk_gp_urls.get("1.0","end").splitlines() if u.strip()]
        if not urls: messagebox.showinfo("Bilgi", "Önce en az bir URL girin."); return
        self._log(f"▶ Manuel GP Yükleme (URL'ler)...")
        for url in urls:
            try:
                resp = http_get(url, token=None, typ="bin", log=self._log, stop_evt=self.stop_evt)
                if resp is None: self._log(f"  ❌ URL yüklenemedi: {url}"); continue
                wb = load_workbook(io.BytesIO(resp.content), data_only=True)
                rows_ready = parse_gp_template_workbook(wb, self._log)
                if rows_ready: self._merge_gp_ready_rows(rows_ready)
                else: self._merge_gp_items(parse_gp_workbook(wb, self._log))
            except Exception as e: self._log(f"  ❌ URL işlenemedi: {e}")
        self._log(f"✅ Manuel URL'den yükleme tamamlandı.")
    def _load_gp_from_file(self):
        p = filedialog.askopenfilename(title="Gider Pusulası Excel seç", filetypes=[("Excel","*.xlsx *.xls")])
        if not p: return
        self._log("▶ Manuel GP Yükleme (Dosya)...")
        try:
            wb = load_workbook(p, data_only=True)
            rows_ready = parse_gp_template_workbook(wb, self._log)
            if rows_ready: self._merge_gp_ready_rows(rows_ready)
            else: self._merge_gp_items(parse_gp_workbook(wb, self._log))
        except Exception as e: messagebox.showerror("Hata", f"Dosya yüklenemedi: {e}")
    def _merge_gp_items(self, items: List[Dict[str,Any]], from_auto_scan=False):
        add_new = self.tk_add_new_imeis.get() if not from_auto_scan else False
        if not items: self._log("[GP] Kayıt bulunamadı."); return
        added = merged = skipped = 0
        for it in items:
            im = it["imei"]
            txt = nup(it.get("aciklama",""))
            self._mark_flags(im, ref=bool(KEY_REF.search(txt)), is2=bool(KEY_2EL.search(txt)), gp=True)
            iid = self.imei_to_iid.get(im)
            brand = brand_from_text(it.get("aciklama","")); model = it.get("aciklama",""); borc  = it.get("bedel",""); sube  = it.get("sube",""); info  = "; ".join([v for v in [("Şube: "+sube) if sube else "", it.get("aciklama","")] if v])
            if iid:
                vals = ensure_len(list(self.tree.item(iid, "values")))
                if not vals[7] and borc: vals[7] = borc
                if vals[17]: vals[17] += (" | " + info) if info else ""
                else: vals[17] = info
                if vals[18] and "Gider Pusulası" not in vals[18]: vals[18] = vals[18] + " + Gider Pusulası"
                elif not vals[18]: vals[18] = "Gider Pusulası"
                if not vals[8] or vals[8]=="Bilinmeyen": vals[8] = brand
                if not vals[9]: vals[9] = model
                if not vals[3]: vals[3] = "GİDER PUSULASI"
                if not vals[19] or vals[19] == "ALIŞ KAYDI GEREKLİ": vals[19] = "Satılabilir"
                vals[1] = "XML+GP" if (vals[1] and vals[1] != "GP") else (vals[1] or "GP")
                self.tree.item(iid, values=vals)
                idx = self.iid_to_row_index.get(iid)
                if idx is not None: self.rows[idx] = ensure_len(vals)
                merged += 1
            else:
                if add_new:
                    if im not in self.force_imeis_set:
                        self.force_imeis_order.append(im); self.force_imeis_set.add(im)
                    row = [ im, "GP", "", "GİDER PUSULASI", it.get("tarih",""), "GP", it.get("ad",""), borc, brand, model, "", "", "", "", "", "", "", info, "Gider Pusulası", "Satılabilir", "", "", "", "" ]
                    iid = self.tree.insert("", "end", values=ensure_len(row))
                    self.iid_to_row_index[iid] = len(self.rows); self.rows.append(ensure_len(row)); self.imei_to_iid.setdefault(im, iid); added += 1
                else: skipped +=1
            if self.imei_to_iid.get(im):
                self._update_classification_for(im)
        log_prefix = "[Otomatik GP]" if from_auto_scan else "[Manuel GP]"
        self._log(f"{log_prefix} Birleştirilen={merged}, Yeni Eklenen={added}, Atlanan={skipped}")
    def _mark_flags(self, imei: str, *, ref: Optional[bool]=None, is2: Optional[bool]=None, gp: Optional[bool]=None):
        f = self.imei_flags.get(imei, {"ref":False,"2el":False,"gp":False})
        if ref is not None: f["ref"] = f["ref"] or ref
        if is2 is not None: f["2el"] = f["2el"] or is2
        if gp  is not None: f["gp"]  = f["gp"]  or gp
        self.imei_flags[imei] = f
    def _add_kdv_in(self, imei: str, kdv: Optional[float]):
        if kdv is None: return
        s = self.imei_kdv_in.get(imei, set()); s.add(int(round(kdv))); self.imei_kdv_in[imei] = s
    def _add_kdv_out(self, imei: str, kdv: Optional[float]):
        if kdv is None: return
        s = self.imei_kdv_out.get(imei, set()); s.add(int(round(kdv))); self.imei_kdv_out[imei] = s
    def _find_kdv_for_imei(self, P: Parsed, imei: str) -> Optional[float]:
        for L in P.lines:
            if imei and imei in L["blob"]: return L.get("kdv", None) or P.inv_kdv
        return P.inv_kdv
    def _stringify_kdvset(self, s: Set[int]) -> str:
        if not s: return ""
        return ",".join([str(x) for x in sorted(s)])
    def _update_kdv_cols(self, imei: str):
        iid = self.imei_to_iid.get(imei)
        if not iid: return
        vals = ensure_len(list(self.tree.item(iid, "values")))
        vals[20] = self._stringify_kdvset(self.imei_kdv_in.get(imei, set()))
        vals[21] = self._stringify_kdvset(self.imei_kdv_out.get(imei, set()))
        self.tree.item(iid, values=vals)
        idx = self.iid_to_row_index.get(iid)
        if idx is not None: self.rows[idx] = ensure_len(vals)
    def _update_classification_for(self, imei: str):
        iid = self.imei_to_iid.get(imei);
        if not iid: return
        vals = ensure_len(list(self.tree.item(iid, "values")))
        k_in  = self.imei_kdv_in.get(imei, set()); k_out = self.imei_kdv_out.get(imei, set()); f = self.imei_flags.get(imei, {"ref":False,"2el":False,"gp":False}); reasons = []; klass = ""
        has_satis_1 = (1 in k_out); has_alis_1  = (1 in k_in); has_alis_20 = (20 in k_in); has_satis_20= (20 in k_out)
        if has_satis_1:
            klass = "YENİLENMİŞ"; reasons.append("SATIŞ KDV=1")
            if has_alis_1: reasons.append("ALIŞ 1 → SATIŞ 1")
            if has_alis_20: reasons.append("ALIŞ 20 → SATIŞ 1")
            if f.get("gp"): reasons.append("Gider pusulası + SATIŞ 1")
        if not klass and f.get("ref"): klass = "YENİLENMİŞ"; reasons.append("Metin ipucu: YENİLENMİŞ/REFURB")
        if not klass and f.get("2el"):
            klass = "2.EL"; reasons.append("Metin ipucu: 2.EL")
            if has_alis_20 and has_satis_20: reasons.append("ALIŞ 20 → SATIŞ 20")
        vals[20] = self._stringify_kdvset(k_in); vals[21] = self._stringify_kdvset(k_out); vals[22] = klass; vals[23] = "; ".join(reasons)
        self.tree.item(iid, values=vals)
        idx = self.iid_to_row_index.get(iid)
        if idx is not None: self.rows[idx] = ensure_len(vals)
    def _start_scan(self):
        if self.worker and self.worker.is_alive(): messagebox.showinfo("Bilgi", "Devam eden iş var. Önce durdurun."); return
        if not self.tk_token.get().strip(): messagebox.showwarning("Uyarı", "Önce API token girin."); return
        if not self.force_imeis_set: messagebox.showwarning("Uyarı", "Lütfen önce 'IMEI Listesi Yükle' ile bir başlangıç listesi seçin."); return
        global TIMEOUT_CONNECT, TIMEOUT_READ, RETRIES, BACKOFF, SESSION
        TIMEOUT_CONNECT = int(self.tk_timeout_c.get() or DEFAULTS["timeout_connect"]); TIMEOUT_READ = int(self.tk_timeout_r.get() or DEFAULTS["timeout_read"]); RETRIES = int(self.tk_retries.get() or DEFAULTS["retries"]); BACKOFF = float(self.tk_backoff.get() or DEFAULTS["backoff"]); SESSION = make_session(RETRIES, BACKOFF)
        terms_text = self.tk_terms.get("1.0","end"); doc_from_terms = self._extract_docnos(terms_text)
        if doc_from_terms:
            self.docnos_filter.update(doc_from_terms)
            self.tk_docnos_count.set(f"Seçili fatura: {len(self.docnos_filter)}")
        self.log.delete("1.0", tk.END); self.stop_evt.clear(); self.btn_stop.config(state="normal")
        self.worker = threading.Thread(target=self._scan_flow, daemon=True); self.worker.start()
    def _scan_flow(self):
        try:
            token = self.tk_token.get().strip(); start = self.tk_start.get().strip() if self.tk_use_date.get() else ""; end   = self.tk_end.get().strip() if self.tk_use_date.get() else ""
            self._log("▶▶▶ Rapor Tamamlama Süreci Başladı...")
            # ADIM 1: NES ARAMASI
            self._log("1. Adım: NES API üzerinden faturalar taranıyor...")
            sales_first = self.tk_scan_sales.get() and bool(self.docnos_filter)
            if sales_first:
                self._log("🔸 [SATIŞ-ÖNCELİKLİ MOD] EAR/EFR listesi yüklü → önce satışlar taranacak.")
                out_arch = list_both_archived(EARCH_OUT_LIST, token, start, end, log=self._log, stop_evt=self.stop_evt, section_name="SATIŞ-eArşiv")
                for meta in out_arch:
                    if self.stop_evt.is_set(): break
                    doc_no_meta = str(meta.get("documentNumber") or "")
                    if self.docnos_filter and doc_no_meta.upper() not in self.docnos_filter: continue
                    inv_id = str(meta.get("id") or ""); xmlb = fetch_xml_by(EARCH_OUT_DOC, token, inv_id)
                    if not xmlb: continue
                    P = parse_invoice_xml(xmlb)
                    if not P.imeis: continue
                    for im in P.imeis: self._append_or_merge_sale(P, inv_id, P.invoice_no or doc_no_meta or inv_id, im, kind="E-ARŞİV")
                out_einv = list_both_archived(EINV_OUT_LIST, token, start, end, log=self._log, stop_evt=self.stop_evt, section_name="SATIŞ-Giden")
                for meta in out_einv:
                    if self.stop_evt.is_set(): break
                    doc_no_meta = str(meta.get("documentNumber") or "")
                    if self.docnos_filter and doc_no_meta.upper() not in self.docnos_filter: continue
                    inv_id = str(meta.get("id") or ""); xmlb = fetch_xml_by(EINV_OUT_DOC, token, inv_id)
                    if not xmlb: continue
                    P = parse_invoice_xml(xmlb)
                    if not P.imeis: continue
                    for im in P.imeis: self._append_or_merge_sale(P, inv_id, P.invoice_no or doc_no_meta or inv_id, im, kind="E-FATURA")
            
            self._log("🔹 [ALIŞ] Tarama başlıyor...")
            incoming = list_both_archived(EINV_IN_LIST, token, start, end, log=self._log, stop_evt=self.stop_evt, section_name="ALIŞ")
            self._log(f"🔹 [ALIŞ] Toplam tekil fatura: {len(incoming)}")
            found_imeis = set(); refurb_rows: List[List[Any]] = []; renewal_rows: List[List[Any]] = []
            for idx, meta in enumerate(incoming, 1):
                if self.stop_evt.is_set(): break
                inv_id = str(meta.get("id") or ""); doc_no = str(meta.get("documentNumber") or inv_id)
                xmlb = fetch_xml_by(EINV_IN_DOC, token, inv_id)
                if not xmlb: self._log(f"[ALIŞ] {idx}/{len(incoming)} {doc_no}: XML indirilemedi."); continue
                P = parse_invoice_xml(xmlb)
                if not P.imeis and is_whitelisted_supplier(P.supplier_name, self.settings): continue
                if self.tk_unvan.get().strip() and (nlow(self.tk_unvan.get()) not in nlow(P.supplier_name)): continue
                if self.tk_tckn.get().strip() and P.supplier_id != self.tk_tckn.get().strip(): continue
                if self.tk_vkn.get().strip() and P.supplier_id != self.tk_vkn.get().strip(): continue
                
                imeis_to_process = [x for x in P.imeis if x in self.force_imeis_set]

                for im in imeis_to_process:
                    unit_price = ""; line_total = ""; model = P.model; brand = P.brand
                    for L in P.lines:
                        if im in L["blob"]: unit_price = L.get("unit_price","") or ""; line_total = L.get("line_total","") or ""; model = L["blob"]; brand = brand_from_text(L["blob"]); break
                    borc_tutar = unit_price or line_total or P.payable
                    self._append_or_merge_purchase(P, inv_id, P.invoice_no or doc_no, im, borc_tutar, model, brand); found_imeis.add(im)
                
                if not P.imeis:
                    tU = P.text_upper
                    if KEY_REF.search(tU): refurb_rows.append(ensure_len(["","XML",P.supplier_id,"FATURA",P.issue_date,P.invoice_no or doc_no,P.supplier_name, P.payable, brand_from_text(tU), P.model or "", "","","","","","","", "YENİLENMİŞ ürün (IMEI yok)","Fatura","Satılabilir", "","","",""]))
                    if "CEP TELEFONU YENİLEME HİZMETİ" in tU: renewal_rows.append(ensure_len(["", "XML", P.supplier_id, "FATURA (Hizmet)", P.issue_date, P.invoice_no or doc_no, P.supplier_name, P.payable, brand_from_text(tU), P.model or "", "", "", "", "", "", "", "", P.description or "CEP TELEFONU YENİLEME HİZMETİ", "Fatura (Hizmet)", "ALIŞ KAYDI GEREKLİ", "", "", "", "Yenileme Faturası, GP'den eşleştirilmeli" ]))
            for r in refurb_rows + renewal_rows:
                iid = self.tree.insert("", "end", values=r); self.iid_to_row_index[iid] = len(self.rows); self.rows.append(r)
            
            if self.tk_scan_sales.get() and not self.stop_evt.is_set() and not sales_first:
                self._log("🔸 [SATIŞ] Tarama başlıyor...")
                out_arch = list_both_archived(EARCH_OUT_LIST, token, start, end, log=self._log, stop_evt=self.stop_evt, section_name="SATIŞ-eArşiv")
                for meta in out_arch:
                    if self.stop_evt.is_set(): break
                    inv_id = str(meta.get("id") or ""); doc_no = str(meta.get("documentNumber") or inv_id); xmlb = fetch_xml_by(EARCH_OUT_DOC, token, inv_id)
                    if not xmlb: continue
                    P = parse_invoice_xml(xmlb)
                    if P.imeis: [self._append_or_merge_sale(P, inv_id, P.invoice_no or doc_no, im, kind="E-ARŞİV") for im in P.imeis]
                out_einv = list_both_archived(EINV_OUT_LIST, token, start, end, log=self._log, stop_evt=self.stop_evt, section_name="SATIŞ-Giden")
                for meta in out_einv:
                    if self.stop_evt.is_set(): break
                    inv_id = str(meta.get("id") or ""); doc_no = str(meta.get("documentNumber") or inv_id); xmlb = fetch_xml_by(EINV_OUT_DOC, token, inv_id)
                    if not xmlb: continue
                    P = parse_invoice_xml(xmlb)
                    if P.imeis: [self._append_or_merge_sale(P, inv_id, P.invoice_no or doc_no, im, kind="E-FATURA") for im in P.imeis]
            
            self._log("✅ 1. Adım (NES Arama) Tamamlandı.")
            
            # ADIM 2: OTOMATİK GP ARAMASI
            urls = [u.strip() for u in self.tk_gp_urls.get("1.0","end").splitlines() if u.strip()]
            if urls:
                self._log("2. Adım: GP Linki ile eksik alış bilgileri tamamlanıyor...")
                for url in urls:
                    try:
                        resp = http_get(url, token=None, typ="bin", log=self._log, stop_evt=self.stop_evt)
                        if resp is None: self._log(f"  ❌ GP URL yüklenemedi: {url}"); continue
                        wb = load_workbook(io.BytesIO(resp.content), data_only=True)
                        items = parse_gp_workbook(wb, self._log)
                        if items: self._merge_gp_items(items, from_auto_scan=True)
                    except Exception as e: self._log(f"  ❌ GP URL işlenemedi: {e}")
                self._log("✅ 2. Adım (Otomatik GP Tamamlama) Bitti.")

            if self.tk_auto_xlsx.get():
                outp = self.settings.get("out_name", DEFAULTS["out_name"])
                try: write_excel(self._table_rows(), outp); self._log(f"🧾 Excel yazıldı: {outp}")
                except Exception as e: self._log(f"❌ Excel yazılamadı: {e}")
            self._log("▶▶▶ Rapor Tamamlama Süreci Bitti.")
        except Exception as e: messagebox.showerror("Hata", str(e))
        finally: self.btn_stop.config(state="disabled")
    def _download_selected(self):
        sel = self.tree.selection()
        if not sel: messagebox.showinfo("Bilgi", "Listeden en az bir satır seçin."); return
        token = self.tk_token.get().strip()
        if not token: messagebox.showwarning("Uyarı", "Önce token girin."); return
        want_pdf = self.tk_get_pdf.get(); want_xml = self.tk_get_xml.get()
        if not (want_pdf or want_xml): messagebox.showinfo("Bilgi", "PDF ve/veya XML işaretleyin."); return
        save_dir = self.tk_dir.get().strip() or os.getcwd(); os.makedirs(save_dir, exist_ok=True); saved = 0
        for iid in sel:
            ids = self.iid_to_ids.get(iid, {}); in_id = ids.get("in_id"); in_doc = safe_filename(ids.get("in_doc","ALIS")); out_id = ids.get("out_id"); out_doc = safe_filename(ids.get("out_doc","SATIS")); out_kind= ids.get("out_kind","E-ARŞİV")
            if in_id:
                if want_xml and (xmlb := fetch_xml_by(EINV_IN_DOC, token, in_id)):
                    with open(os.path.join(save_dir, f"{in_doc}_ALIS.xml"), "wb") as f: f.write(xmlb); saved += 1
                if want_pdf and (pdfb := fetch_pdf_by(EINV_IN_DOC, token, in_id)):
                    with open(os.path.join(save_dir, f"{in_doc}_ALIS.pdf"), "wb") as f: f.write(pdfb); saved += 1
            if out_id:
                DOC = EARCH_OUT_DOC if out_kind=="E-ARŞİV" else EINV_OUT_DOC
                if want_xml and (xmlb := fetch_xml_by(DOC, token, out_id)):
                    with open(os.path.join(save_dir, f"{out_doc}_SATIS.xml"), "wb") as f: f.write(xmlb); saved += 1
                if want_pdf and (pdfb := fetch_pdf_by(DOC, token, out_id)):
                    with open(os.path.join(save_dir, f"{out_doc}_SATIS.pdf"), "wb") as f: f.write(pdfb); saved += 1
        self._log(f"📦 İndirme tamamlandı. Kaydedilen dosya: {saved}")
        if saved:
            try: os.startfile(save_dir)
            except Exception: pass
    def _table_rows(self) -> List[List[Any]]:
        return [ensure_len(list(self.tree.item(iid, "values"))) for iid in self.tree.get_children("")]
    def _export_excel(self):
        rows = self._table_rows()
        if not rows: messagebox.showinfo("Bilgi", "Henüz sonuç yok."); return
        p = filedialog.asksaveasfilename(title="Excel kaydet", defaultextension=".xlsx", initialfile=os.path.basename(self.settings.get("out_name", DEFAULTS["out_name"])), filetypes=[("Excel", "*.xlsx")])
        if not p: return
        try: write_excel(rows, p); self._log(f"🧾 Excel yazıldı: {p}")
        except Exception as e: messagebox.showerror("Hata", f"Excel yazılamadı: {e}")
    def _append_or_merge_purchase(self, P: Parsed, inv_id: str, doc_no: str, imei: str, payable: str, model: str, brand: str):
        if imei not in self.force_imeis_set: return # Sadece ana listedeki IMEI'leri işle
        pair = (imei, doc_no)
        if pair in self.seen_in_pairs: return
        self.seen_in_pairs.add(pair)
        self._mark_flags(imei, ref=bool(KEY_REF.search(P.text_upper)), is2=bool(KEY_2EL.search(P.text_upper)))
        self._add_kdv_in(imei, self._find_kdv_for_imei(P, imei))
        iid = self.imei_to_iid.get(imei)
        if iid:
            vals = ensure_len(list(self.tree.item(iid, "values")))
            if not vals[4]: vals[4] = P.issue_date
            if not vals[5]: vals[5] = doc_no
            if not vals[6]: vals[6] = P.supplier_name
            if not vals[7]: vals[7] = payable
            if not vals[8] or vals[8]=="Bilinmeyen": vals[8] = brand
            if not vals[9]: vals[9] = model
            if not vals[2]: vals[2] = P.supplier_id
            if not vals[3]: vals[3] = "FATURA"
            if vals[18] and "Fatura" not in vals[18]: vals[18] += " + Fatura"
            elif not vals[18]: vals[18] = "Fatura"
            vals[1] = "XML" if not vals[1] or vals[1]=="Bulunamadı" else vals[1]
            if vals[19] == "ALIŞ KAYDI GEREKLİ": vals[19] = "Satılabilir"
            self.tree.item(iid, values=vals)
            idx = self.iid_to_row_index.get(iid)
            if idx is not None: self.rows[idx] = ensure_len(vals)
            ids = self.iid_to_ids.setdefault(iid, {})
            if not ids.get("in_id"): ids["in_id"] = inv_id; ids["in_doc"] = doc_no
            self._update_kdv_cols(imei); self._update_classification_for(imei)
    def _append_or_merge_sale(self, P: Parsed, inv_id: str, doc_no: str, imei: str, kind: str):
        if imei not in self.force_imeis_set: return # Sadece ana listedeki IMEI'leri işle
        pair = (imei, doc_no)
        if pair in self.seen_out_pairs: return
        self.seen_out_pairs.add(pair)
        self._mark_flags(imei, ref=bool(KEY_REF.search(P.text_upper)), is2=bool(KEY_2EL.search(P.text_upper)))
        self._add_kdv_out(imei, self._find_kdv_for_imei(P, imei))
        iid = self.imei_to_iid.get(imei)
        if iid:
            vals = ensure_len(list(self.tree.item(iid, "values")))
            if not vals[10]: vals[10] = P.issue_date
            if not vals[11]: vals[11] = P.buyer_name
            if not vals[12]: vals[12] = P.payable
            if not vals[13]: vals[13] = P.tax_total
            if not vals[14]: vals[14] = P.invoice_no or doc_no
            if not vals[15]: vals[15] = P.buyer_id_type
            if not vals[16]: vals[16] = P.buyer_id
            vals[19] = "Satılmış"
            self.tree.item(iid, values=vals)
            idx = self.iid_to_row_index.get(iid)
            if idx is not None: self.rows[idx] = ensure_len(vals)
            ids = self.iid_to_ids.setdefault(iid, {})
            if not ids.get("out_id"): ids["out_id"] = inv_id; ids["out_doc"] = doc_no; ids["out_kind"] = kind
            self._update_kdv_cols(imei); self._update_classification_for(imei)

if __name__ == "__main__":
    App().mainloop()
